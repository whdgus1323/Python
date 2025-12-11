import re
import time
import os
import sys
import threading
import queue
import requests
import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, font as tkfont

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import load_workbook
from openpyxl.styles import Font
from bs4 import BeautifulSoup

USERNAME = "tb548563353414"
PASSWORD = "zxc@73026181"

LOGIN_URL = "https://login.1688.com/member/signin.htm"
SEARCH_URL_TEMPLATE = "https://s.1688.com/selloffer/offer_search.htm?keywords={keyword}"

PAPAGO_URL = "https://papago.naver.com/apis/n2mt/translate"
PAPAGO_HEADERS = {
    "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
    "Accept": "*/*",
    "User-Agent": "Mozilla/5.0",
}

FX_URL = "https://open.er-api.com/v6/latest/CNY"

log_widget = None
log_queue = queue.Queue()
result_queue = queue.Queue()

if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

ICON_ICO = os.path.join(BASE_DIR, "app_icon.ico")
ICON_PNG = os.path.join(BASE_DIR, "app_icon.png")
LAST_DIR_FILE = os.path.join(BASE_DIR, "last_save_dir.txt")


def log(msg: str):
    text = str(msg)
    print(text)
    try:
        log_queue.put_nowait(text)
    except queue.Full:
        pass


def clean_price_from_card_text(text: str) -> str:
    if not isinstance(text, str):
        return ""
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    if "¥" in lines:
        idx = lines.index("¥")
        parts = []
        for t in lines[idx + 1: idx + 6]:
            if re.match(r"^[0-9.]+$", t):
                parts.append(t)
            else:
                break
        if parts:
            price = "¥" + "".join(parts)
            return price.replace(" ", "")
    m = re.search(r"¥\s*([0-9.]+)", text)
    if m:
        return "¥" + m.group(1)
    return ""


def extract_offer_id(url: str) -> str:
    if not isinstance(url, str):
        return ""
    m = re.search(r"offerId=(\d+)", url)
    if m:
        return m.group(1)
    m = re.search(r"/offer/(\d+)\.html", url)
    if m:
        return m.group(1)
    return ""


def create_driver():
    log("[DEBUG] Edge 드라이버 생성")
    options = webdriver.EdgeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Edge(options=options)
    driver.implicitly_wait(5)
    return driver


def login_1688(driver):
    log(f"[DEBUG] 로그인 페이지 접속: {LOGIN_URL}")
    driver.get(LOGIN_URL)
    wait = WebDriverWait(driver, 20)
    id_input = wait.until(EC.presence_of_element_located((By.XPATH, "//input[contains(@placeholder,'账号名') or contains(@placeholder,'邮箱') or contains(@placeholder,'手机号')]")))
    id_input.clear()
    id_input.send_keys(USERNAME)
    log("[DEBUG] 아이디 입력 완료")
    pw_input = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='password']")))
    pw_input.clear()
    pw_input.send_keys(PASSWORD)
    log("[DEBUG] 비밀번호 입력 완료")
    login_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(),'登录')]")))
    login_btn.click()
    log("[DEBUG] 로그인 버튼 클릭")
    time.sleep(5)
    log(f"[DEBUG] 로그인 후 URL: {driver.current_url}")
    log(f"[DEBUG] 로그인 후 TITLE: {driver.title}")


def go_search_page(driver, keyword):
    from urllib.parse import quote_plus
    encoded = quote_plus(keyword, encoding="gbk", errors="ignore")
    url = SEARCH_URL_TEMPLATE.format(keyword=encoded)
    log(f"[DEBUG] 검색 URL: {url}")
    driver.get(url)
    time.sleep(5)
    log(f"[DEBUG] 검색 페이지 TITLE: {driver.title}")


def scroll_page(driver, times=3, pause=2.0):
    log(f"[DEBUG] 스크롤 {times}회 시작")
    for i in range(times):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(pause)
        log(f"[DEBUG] 스크롤 {i + 1}/{times} 완료")
    log("[DEBUG] 스크롤 완료")


def collect_cards(driver):
    selector = "a.search-offer-wrapper"
    log(f"[DEBUG] 상품 카드 셀렉터: {selector}")
    cards = driver.find_elements(By.CSS_SELECTOR, selector)
    log(f"[DEBUG] 찾은 상품 카드 개수: {len(cards)}")
    products = []
    for card in cards:
        raw = card.text.strip()
        href = card.get_attribute("href")
        offer_id = extract_offer_id(href)
        name_raw = raw.split("\n")[0] if raw else ""
        price_raw = clean_price_from_card_text(raw)
        products.append({"name_raw": name_raw, "price_raw": price_raw, "url": href, "offer_id": offer_id})
    return products


def build_requests_session(driver):
    log("[DEBUG] Selenium 쿠키 → requests 세션 복사")
    s = requests.Session()
    for c in driver.get_cookies():
        s.cookies.set(c["name"], c["value"])
    s.headers.update({"User-Agent": "Mozilla/5.0", "Accept": "*/*"})
    return s


def get_cny_to_krw_rate():
    log(f"[DEBUG] 환율 요청: {FX_URL}")
    resp = requests.get(FX_URL, timeout=10)
    resp.raise_for_status()
    data = resp.json()
    if data.get("result") != "success":
        raise RuntimeError(f"환율 API 오류: {data}")
    rate = data.get("rates", {}).get("KRW")
    if rate is None:
        raise RuntimeError("KRW 환율을 가져오지 못했습니다.")
    log(f"[DEBUG] CNY→KRW 환율: {rate}")
    return float(rate)


def papago_translate(text: str, session: requests.Session) -> str:
    if not isinstance(text, str) or not text.strip():
        return ""
    data = {"source": "zh-CN", "target": "ko", "text": text}
    try:
        r = session.post(PAPAGO_URL, headers=PAPAGO_HEADERS, data=data, timeout=10)
        r.raise_for_status()
        j = r.json()
        translated = j.get("translatedText", "")
        if not translated:
            log("[WARN] Papago 응답에 translatedText 없음, 원문 사용")
            return text
        return translated
    except Exception as e:
        log(f"[WARN] Papago 번역 실패, 원문 사용: {e}")
        return text


def translate_list(text_list, session):
    result = []
    total = len(text_list)
    for i, t in enumerate(text_list, start=1):
        log(f"[DEBUG] 번역 {i}/{total}…")
        translated = papago_translate(t, session)
        result.append(translated)
        time.sleep(0.4)
    return result


def parse_price_cny(price_raw: str) -> float:
    if not isinstance(price_raw, str):
        return 0.0
    m = re.search(r"([0-9]+(?:\.[0-9]+)?)", price_raw)
    if not m:
        return 0.0
    try:
        return float(m.group(1))
    except ValueError:
        return 0.0


def fetch_detail_attributes(session: requests.Session, url: str) -> dict:
    try:
        log(f"[DEBUG] 상세페이지 요청: {url}")
        r = session.get(url, timeout=15)
        r.raise_for_status()
        html = r.text
    except Exception as e:
        log(f"[WARN] 상세페이지 요청 실패: {e}")
        return {}
    try:
        soup = BeautifulSoup(html, "html.parser")
    except Exception as e:
        log(f"[WARN] 상세페이지 파싱 실패: {e}")
        return {}
    heading = soup.find(string=lambda t: isinstance(t, str) and "제품 속성" in t)
    if not heading:
        return {}
    parent = heading.find_parent()
    if not parent:
        return {}
    table = parent.find_next("table")
    if not table:
        return {}
    attrs = {}
    for tr in table.find_all("tr"):
        cells = tr.find_all(["th", "td"])
        if len(cells) < 2:
            continue
        for i in (0, 2):
            if i >= len(cells) - 1:
                continue
            key = cells[i].get_text(" ", strip=True).strip().rstrip(":")
            val = cells[i + 1].get_text(" ", strip=True)
            if key:
                attrs[key] = val
    return attrs


def run_crawler(keywords, scroll_times, output_file):
    driver = create_driver()
    all_products = []
    try:
        login_1688(driver)
        for kw in keywords:
            log(f"[DEBUG] 검색어 처리: {kw}")
            go_search_page(driver, kw)
            scroll_page(driver, times=scroll_times, pause=2.0)
            products = collect_cards(driver)
            for p in products:
                p["keyword"] = kw
            all_products.extend(products)
        if not all_products:
            log("[ERROR] 카드가 하나도 없습니다.")
            return False
        http_session = build_requests_session(driver)
    finally:
        try:
            driver.quit()
        except:
            pass

    rate_cny_krw = get_cny_to_krw_rate()
    names_cn = [p["name_raw"] for p in all_products]
    names_ko = translate_list(names_cn, http_session)

    rows = []
    total = len(all_products)
    for idx, (p, name_ko) in enumerate(zip(all_products, names_ko), start=1):
        log(f"[DEBUG] 상세정보 포함 행 생성 {idx}/{total}")
        price_cny = parse_price_cny(p["price_raw"])
        price_krw = int(round(price_cny * rate_cny_krw))
        detail_attrs = fetch_detail_attributes(http_session, p["url"])
        row = {
            "검색어": p.get("keyword", ""),
            "상품명": name_ko,
            "공급가(원)": price_krw,
            "링크": p["url"],
        }
        row.update(detail_attrs)
        rows.append(row)
        time.sleep(0.5)

    df = pd.DataFrame(rows)
    df.to_excel(output_file, index=False)
    log(f"[DEBUG] 1차 저장 완료: {output_file}")

    wb = load_workbook(output_file)
    ws = wb.active
    cols = list(df.columns)
    if "링크" in cols:
        link_col_idx = cols.index("링크") + 1
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=link_col_idx)
            link = cell.value
            if link:
                cell.hyperlink = link
                cell.font = Font(color="0000FF", underline="single")
    wb.save(output_file)
    log(f"[DEBUG] 최종 저장 완료: {output_file}")
    return True


def run_crawler_thread(keywords, scroll_times, output_file, dir_path):
    try:
        ok = run_crawler(keywords, scroll_times, output_file)
        if not ok:
            result_queue.put(("error", "카드가 하나도 없습니다.", output_file, dir_path))
        else:
            result_queue.put(("ok", "", output_file, dir_path))
    except Exception as e:
        result_queue.put(("error", str(e), output_file, dir_path))


def on_browse_dir():
    initial = save_dir_var.get().strip() or os.getcwd()
    path = filedialog.askdirectory(initialdir=initial)
    if path:
        save_dir_var.set(path)


def load_last_dir():
    if os.path.exists(LAST_DIR_FILE):
        try:
            with open(LAST_DIR_FILE, "r", encoding="utf-8") as f:
                d = f.read().strip()
                if d:
                    save_dir_var.set(d)
        except Exception:
            pass


def save_last_dir(path):
    try:
        with open(LAST_DIR_FILE, "w", encoding="utf-8") as f:
            f.write(path)
    except Exception:
        pass


def paste_keywords(event=None):
    try:
        text = root.clipboard_get()
    except tk.TclError:
        return "break"
    entry_keywords.delete(0, tk.END)
    entry_keywords.insert(0, text)
    log(f"[DEBUG] 붙여넣기 텍스트: {repr(text)}")
    return "break"


def on_run():
    raw_keywords = entry_keywords.get().strip()
    log(f"[DEBUG] on_run 키워드(raw): {repr(raw_keywords)}")
    scroll_str = entry_scroll.get().strip()
    if not raw_keywords:
        messagebox.showerror("오류", "상품 키워드를 입력하세요. 예: glass,cup,bottle")
        return
    try:
        scroll_times = int(scroll_str) if scroll_str else 3
        if scroll_times <= 0:
            raise ValueError
    except ValueError:
        messagebox.showerror("오류", "스크롤 횟수는 1 이상의 정수여야 합니다.")
        return
    keywords = [k.strip() for k in raw_keywords.split(",") if k.strip()]
    if not keywords:
        messagebox.showerror("오류", "유효한 상품 키워드를 입력하세요.")
        return

    dir_path = save_dir_var.get().strip()
    if not dir_path:
        dir_path = os.getcwd()
    filename = entry_filename.get().strip()
    if not filename:
        filename = "1688_products_korean.xlsx"
    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"
    output_file = os.path.join(dir_path, filename)

    log(f"[DEBUG] 출력 파일 경로: {output_file}")

    btn_run.config(state="disabled")
    t = threading.Thread(target=run_crawler_thread, args=(keywords, scroll_times, output_file, dir_path), daemon=True)
    t.start()


def process_log_queue():
    try:
        while True:
            msg = log_queue.get_nowait()
            if log_widget is not None:
                log_widget.insert("end", msg + "\n")
                log_widget.see("end")
    except queue.Empty:
        pass
    root.after(100, process_log_queue)


def process_result_queue():
    try:
        while True:
            kind, err_msg, output_file, dir_path = result_queue.get_nowait()
            btn_run.config(state="normal")
            if kind == "ok":
                if remember_var.get():
                    save_last_dir(dir_path)
                messagebox.showinfo("완료", f"작업이 완료되었습니다.\n출력 파일: {output_file}")
            else:
                log(f"[ERROR] 작업 중 오류 발생: {err_msg}")
                messagebox.showerror("오류", f"작업 중 오류 발생:\n{err_msg}")
    except queue.Empty:
        pass
    root.after(200, process_result_queue)


root = tk.Tk()
root.title("1688 크롤러")

try:
    root.iconbitmap(ICON_ICO)
except:
    try:
        icon_img = tk.PhotoImage(file=ICON_PNG)
        root.iconphoto(True, icon_img)
        root._icon_img = icon_img
    except:
        pass

root.geometry("1100x600")
root.minsize(900, 500)
root.configure(bg="#111827")

style = ttk.Style()
try:
    style.theme_use("clam")
except:
    pass

font_family_main = "Microsoft YaHei UI"

default_font = tkfont.nametofont("TkDefaultFont")
try:
    default_font.configure(family=font_family_main, size=10)
except tk.TclError:
    default_font.configure(size=10)
root.option_add("*Font", default_font)

style.configure("Main.TFrame", background="#111827")
style.configure("Card.TFrame", background="#1F2933", borderwidth=0, relief="flat")
style.configure("TLabel", background="#1F2933", foreground="#E5E7EB")
style.configure("Header.TLabel", background="#111827", foreground="#F9FAFB", font=(font_family_main, 14))
style.configure("Accent.TButton", background="#3B82F6", foreground="#FFFFFF", borderwidth=0, focusthickness=0, padding=(12, 6))
style.map("Accent.TButton", background=[("active", "#2563EB"), ("disabled", "#4B5563")])
style.configure("TEntry", fieldbackground="#111827", foreground="#F9FAFB")
style.configure("TCheckbutton", background="#1F2933", foreground="#E5E7EB")

main_frame = ttk.Frame(root, padding=10, style="Main.TFrame")
main_frame.grid(row=0, column=0, sticky="nsew")

root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)

main_frame.columnconfigure(0, weight=0)
main_frame.columnconfigure(1, weight=1)
main_frame.rowconfigure(1, weight=1)

header_frame = ttk.Frame(main_frame, style="Main.TFrame")
header_frame.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 8))
header_frame.columnconfigure(0, weight=1)

title_label = ttk.Label(header_frame, text="1688 상품 크롤링", style="Header.TLabel")
title_label.grid(row=0, column=0, sticky="w")

subtitle = ttk.Label(header_frame, text="키워드별 상품 정보 크롤링 후 Papago로 번역하고, 원화 기준 공급가를 계산 (로그인 시 슬라이드 수동)", style="Header.TLabel", font=(font_family_main, 9))
subtitle.configure(foreground="#9CA3AF", background="#111827")
subtitle.grid(row=1, column=0, sticky="w", pady=(2, 0))

left_frame = ttk.Frame(main_frame, padding=16, style="Card.TFrame")
left_frame.grid(row=1, column=0, sticky="nsew", padx=(0, 8))

right_frame = ttk.Frame(main_frame, padding=16, style="Card.TFrame")
right_frame.grid(row=1, column=1, sticky="nsew")

for i in range(6):
    left_frame.rowconfigure(i, weight=0)
left_frame.rowconfigure(6, weight=1)
left_frame.columnconfigure(1, weight=1)

label_keywords = ttk.Label(left_frame, text="상품 키워드(,로 구분):")
label_keywords.grid(row=0, column=0, sticky="w", pady=(0, 4))
entry_keywords = ttk.Entry(left_frame, width=40)
entry_keywords.grid(row=0, column=1, padx=5, pady=(0, 4), sticky="ew")
entry_keywords.bind("<Control-v>", paste_keywords)
entry_keywords.bind("<Control-V>", paste_keywords)
entry_keywords.configure(font=(font_family_main, 10))

label_scroll = ttk.Label(left_frame, text="스크롤 횟수:")
label_scroll.grid(row=1, column=0, sticky="w", pady=(4, 4))
entry_scroll = ttk.Entry(left_frame, width=10)
entry_scroll.insert(0, "3")
entry_scroll.grid(row=1, column=1, sticky="w", padx=5, pady=(4, 4))

label_dir = ttk.Label(left_frame, text="저장 위치:")
label_dir.grid(row=2, column=0, sticky="w", pady=(4, 4))

save_dir_var = tk.StringVar()
entry_dir = ttk.Entry(left_frame, textvariable=save_dir_var, width=40)
entry_dir.grid(row=2, column=1, padx=5, pady=(4, 4), sticky="ew")

btn_browse = ttk.Button(left_frame, text="폴더 선택", command=on_browse_dir, style="Accent.TButton")
btn_browse.grid(row=2, column=2, padx=(6, 0), pady=(4, 4))

label_filename = ttk.Label(left_frame, text="파일 이름:")
label_filename.grid(row=3, column=0, sticky="w", pady=(4, 4))

entry_filename = ttk.Entry(left_frame, width=40)
entry_filename.insert(0, "1688_product_.xlsx")
entry_filename.grid(row=3, column=1, padx=5, pady=(4, 4), sticky="ew")

remember_var = tk.BooleanVar(value=True)
chk_remember = ttk.Checkbutton(left_frame, text="이전 저장 위치 기억", variable=remember_var)
chk_remember.grid(row=4, column=0, columnspan=2, sticky="w", pady=(8, 4))

btn_run = ttk.Button(left_frame, text="실행", command=on_run, style="Accent.TButton")
btn_run.grid(row=5, column=0, columnspan=3, pady=(12, 0), sticky="ew")

right_frame.rowconfigure(1, weight=1)
right_frame.columnconfigure(0, weight=1)

log_title = ttk.Label(right_frame, text="DEBUG 로그", font=(font_family_main, 10, "bold"))
log_title.grid(row=0, column=0, sticky="w", pady=(0, 4))

log_widget = tk.Text(right_frame, wrap="none", width=60, height=25, bg="#020617", fg="#E5E7EB", insertbackground="#E5E7EB", relief="flat", borderwidth=0)
log_widget.configure(font=(font_family_main, 10))
log_widget.grid(row=1, column=0, sticky="nsew")

scrollbar_y = ttk.Scrollbar(right_frame, orient="vertical", command=log_widget.yview)
scrollbar_y.grid(row=1, column=1, sticky="ns", padx=(6, 0))
log_widget.configure(yscrollcommand=scrollbar_y.set)

scrollbar_x = ttk.Scrollbar(right_frame, orient="horizontal", command=log_widget.xview)
scrollbar_x.grid(row=2, column=0, sticky="ew", pady=(6, 0))
log_widget.configure(xscrollcommand=scrollbar_x.set)

load_last_dir()
process_log_queue()
process_result_queue()

root.mainloop()
