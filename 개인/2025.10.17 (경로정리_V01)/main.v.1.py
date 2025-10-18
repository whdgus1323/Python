# -*- coding: utf-8 -*-
"""
Backup 트리 (Collapsed + Depth Columns) - V14
요청 반영:
- 폴더명이 '.h' 로 끝나면 그 폴더는 **하위 탐색을 중단**하고,
  이름을 '원래이름( .h 제거) + " [Hiddle]"' 로 바꿔 **엑셀/텍스트 모두**에 한 줄만 표시.
  예)  Backup_2024.05.11 (AODV.Sec.ana.c.v.1).h
       →  Backup_2024.05.11 (AODV.Sec.ana.c.v.1) [Hiddle]
- 그 외(박스문자 트리, 열/행 간격 설정)는 V13과 동일.
"""

import os, re, math
from datetime import datetime
from typing import List, Dict, Any, Tuple, Set
import pandas as pd

# ===== 사용자 설정 =====
root_dir = r"D:\연구실"   # 시작 폴더
save_basename = f"_BackupTree_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
excel_path = os.path.join(root_dir, save_basename + ".xlsx")
txt_collapsed_path = os.path.join(root_dir, save_basename + "_collapsed.txt")

# 무시 폴더/파일
IGNORE_DIR_KEYWORDS = {
    "__pycache__", ".git", ".venv", "node_modules",
    "$recycle.bin", "system volume information"
}
IGNORE_IMAGE_EXTS = {
    ".jpg",".jpeg",".png",".bmp",".gif",".tif",".tiff",
    ".webp",".heic",".heif",".raw",".cr2",".nef",".arw"
}
IGNORE_MISC_FILE_EXTS = {".tmp",".temp",".log~"}
IGNORE_FILE_EXTS = IGNORE_IMAGE_EXTS | IGNORE_MISC_FILE_EXTS

# 메인 루트 인식
MAIN_ROOT_PATTERN = re.compile(r"^Backup_\d{4}\.\d{2}\.\d{2}(\b|[^\\/]*)", re.IGNORECASE)

# 이름 정규화
CASE_SENSITIVE = False
SQUASH_SPACES = True

# 출력 제한(None=무제한)
ROW_LIMIT = None

# Excel 표시 파라미터
FONT_NAME = "Consolas"
FONT_SIZE = 11
EXCEL_MIN_COL_WIDTH = 10
EXCEL_MAX_COL_WIDTH = 12
DEFAULT_LEVEL_WIDTH = 14
LEVEL0_EXTRA_WIDTH = 12

# ===== 유틸 =====
def human_size(nbytes: int) -> str:
    if not nbytes: return "0 B"
    units = ["B","KB","MB","GB","TB","PB"]
    i = min(int(math.log(nbytes, 1024)), len(units)-1)
    return f"{round(nbytes/(1024**i), 2)} {units[i]}"

def safe_stat(path: str):
    try: return os.stat(path)
    except Exception: return None

def is_ignored_dir(name: str) -> bool:
    low = name.lower()
    return any(k in low for k in IGNORE_DIR_KEYWORDS)

def norm_ext(name: str) -> str:
    return os.path.splitext(name)[1].lower()

def normalize_filename(name: str) -> str:
    s = name.strip()
    if SQUASH_SPACES:
        s = re.sub(r"\s+", " ", s)
    if not CASE_SENSITIVE:
        s = s.lower()
    return s

def find_main_roots(root: str):
    roots = []
    try:
        for e in os.scandir(root):
            if e.is_dir(follow_symlinks=False) and MAIN_ROOT_PATTERN.match(e.name):
                roots.append(e.path)
    except Exception:
        pass
    roots.sort(key=lambda p: os.path.basename(p))
    return roots or [root]

def compress_numbers(nums: List[int]) -> str:
    if not nums: return ""
    nums = sorted(set(nums))
    parts = []
    start = prev = nums[0]
    for n in nums[1:]:
        if n == prev + 1:
            prev = n
        else:
            parts.append(str(start) if start==prev else f"{start}..{prev}")
            start = prev = n
    parts.append(str(start) if start==prev else f"{start}..{prev}")
    return ",".join(parts)

def split_prefix_number(name: str) -> Tuple[str, int]:
    m = re.match(r"^([^\d]*?)(\d+)$", name)
    if not m: return None, None
    return m.group(1), int(m.group(2))

# ----- .h 폴더 처리 -----
def is_hiddle_folder(name: str) -> bool:
    return name.endswith(".h")

def hiddle_label(name: str) -> str:
    """'.h' 제거 후 ' [Hiddle]' 접미사 추가"""
    return (name[:-2] + " [Hiddle]") if is_hiddle_folder(name) else name

# ===== 1) 파일 인벤토리 =====
def walk_files(root: str) -> List[Dict[str, Any]]:
    """파일만 수집 ('.h' 폴더는 **하위 스캔하지 않음**)"""
    rows: List[Dict[str, Any]] = []
    def dfs(path: str):
        base = os.path.basename(path)
        # .h 폴더면 하위 무시
        if is_hiddle_folder(base):
            return
        try:
            children = sorted(os.scandir(path), key=lambda e: (not e.is_dir(), e.name.lower()))
        except Exception:
            return
        for ch in children:
            if ch.is_dir(follow_symlinks=False):
                if is_ignored_dir(ch.name):
                    continue
                # .h 폴더는 들어가지 않음
                if is_hiddle_folder(ch.name):
                    continue
                dfs(ch.path)
            else:
                ext = norm_ext(ch.name)
                if ext in IGNORE_FILE_EXTS:
                    continue
                st = safe_stat(ch.path)
                rows.append({
                    "Parent": path,
                    "Name": ch.name,
                    "NameNorm": normalize_filename(ch.name),
                    "Path": ch.path,
                    "Size(Bytes)": st.st_size if st else 0,
                    "Size(Human)": human_size(st.st_size) if st else ""
                })
    dfs(root)
    return rows

# ===== 2) 그룹 탐지 =====
def detect_groups(root: str):
    file_rows = walk_files(root)
    if not file_rows:
        return pd.DataFrame(), pd.DataFrame()

    df_files = pd.DataFrame(file_rows)
    df_files["MainRoot"] = os.path.basename(root)

    units = (
        df_files.groupby(["MainRoot","Parent"])["NameNorm"]
        .apply(lambda s: tuple(sorted(set(s)))).reset_index(name="FileSet")
    )
    units["FilesPerUnit"] = units["FileSet"].apply(len)
    units["UnitName"] = units["Parent"].apply(lambda p: os.path.basename(p))
    units[["Prefix","Number"]] = units["UnitName"].apply(lambda n: pd.Series(split_prefix_number(n)))
    units["ParentOfUnit"] = units["Parent"].apply(os.path.dirname)

    grouped = (
        units.groupby(["MainRoot","ParentOfUnit","FileSet","Prefix"], dropna=False)
        .agg(
            UnitCount=("UnitName","nunique"),
            FilesPerUnit=("FilesPerUnit","first"),
            UnitNames=("UnitName", lambda x: sorted(set(x))),
            Numbers=("Number", lambda x: sorted(int(n) for n in x if pd.notna(n))),
            Parents=("Parent", lambda x: sorted(set(x)))
        ).reset_index()
    )
    def units_label(prefix, numbers, names):
        if pd.notna(prefix) and numbers:
            return f"{prefix}({compress_numbers(numbers)})"
        if len(names) <= 10:
            return ", ".join(names)
        return ", ".join(names[:10]) + f" … (+{len(names)-10})"
    grouped["UnitsLabel"] = grouped.apply(lambda r: units_label(r["Prefix"], r["Numbers"], r["UnitNames"]), axis=1)
    grouped["FileList"] = grouped["FileSet"].apply(lambda t: ", ".join(t))
    grouped = grouped[grouped["UnitCount"] >= 2].copy()

    grouped = grouped.sort_values(["ParentOfUnit","UnitCount","FilesPerUnit"], ascending=[True, False, False])
    grouped["PatternID"] = grouped.groupby("MainRoot").cumcount().add(1).astype(str).radd(grouped["MainRoot"] + " :: P")
    return grouped.reset_index(drop=True), units.reset_index(drop=True)

# ===== 3) Collapsed Tree =====
def make_collapsed_rows(root: str, grouped: pd.DataFrame, units: pd.DataFrame) -> List[Dict[str, Any]]:
    """
    '.h' 폴더:
      - 현재 위치에 '이름(.h 제거) + " [Hiddle]"' 라벨로 **한 줄만 추가**
      - 그 하위는 **탐색하지 않음**
    """
    rows: List[Dict[str, Any]] = []
    if grouped is None or not isinstance(grouped, pd.DataFrame) or grouped.empty:
        grouped = pd.DataFrame()

    parent_to_groups: Dict[str, List[dict]] = {}
    if not grouped.empty:
        for _, g in grouped.iterrows():
            parent_to_groups.setdefault(g["ParentOfUnit"], []).append(g)

    grouped_member_paths: Set[str] = set()
    if not grouped.empty:
        for parents in grouped["Parents"]:
            grouped_member_paths.update(parents)

    def dfs(path: str, parts: List[str]):
        base = os.path.basename(path)
        # 현재 폴더 라벨( .h → [Hiddle] )
        cur_parts = parts[:]
        cur_parts[-1] = hiddle_label(cur_parts[-1])
        rows.append({"Parts": cur_parts, "Kind": "Folder", "Name": cur_parts[-1], "Extra": ""})
        if ROW_LIMIT and len(rows) >= ROW_LIMIT:
            return

        # 현재가 .h 폴더면 하위 탐색 중단
        if is_hiddle_folder(base):
            return

        groups = parent_to_groups.get(path, [])
        if groups:
            for g in groups:
                label = f"[Group] {g['UnitsLabel']}  [{g['UnitCount']} folders, {g['FilesPerUnit']} files/unit]"
                rows.append({
                    "Parts": cur_parts[:] + [label],
                    "Kind": "Group",
                    "Name": g["UnitsLabel"],
                    "Extra": g["FileList"]
                })
                if ROW_LIMIT and len(rows) >= ROW_LIMIT:
                    return

        try:
            children = sorted(os.scandir(path), key=lambda e: (not e.is_dir(), e.name.lower()))
        except Exception:
            return
        for ch in children:
            if ch.is_dir(follow_symlinks=False):
                if is_ignored_dir(ch.name):
                    continue
                # 자식이 .h 폴더면 **한 줄만** 추가하고 재귀하지 않음
                if is_hiddle_folder(ch.name):
                    label = hiddle_label(ch.name)
                    rows.append({"Parts": cur_parts[:] + [label], "Kind": "Folder", "Name": label, "Extra": ""})
                    if ROW_LIMIT and len(rows) >= ROW_LIMIT:
                        return
                    continue
                if ch.path in grouped_member_paths:
                    continue
                dfs(ch.path, cur_parts + [ch.name])
                if ROW_LIMIT and len(rows) >= ROW_LIMIT:
                    return
            else:
                if path in parent_to_groups or path in grouped_member_paths:
                    continue
                ext = norm_ext(ch.name)
                if ext in IGNORE_FILE_EXTS:
                    continue
                st = safe_stat(ch.path)
                rows.append({
                    "Parts": cur_parts[:] + [ch.name],
                    "Kind": "File",
                    "Name": ch.name,
                    "Extra": human_size(st.st_size) if st else ""
                })
                if ROW_LIMIT and len(rows) >= ROW_LIMIT:
                    return

    # 시작(루트도 .h 규칙 적용)
    dfs(root, [os.path.basename(root)])
    return rows

# ===== 4) 박스문자 트리 매트릭스(├──/└──/│) =====
def build_box_matrix(df_paths: pd.DataFrame, level_cols: List[str]) -> pd.DataFrame:
    paths = [[r[c] for c in level_cols if r[c]] for _, r in df_paths.iterrows()]
    out = []
    for i, r in df_paths.iterrows():
        parts = paths[i]
        vis = {c: "" for c in level_cols}
        if not parts:
            out.append({**vis, "Kind": r["Kind"], "MainRoot": r["MainRoot"], "Extra": r["Extra"]})
            continue
        k = len(parts) - 1
        vis[level_cols[k]] = parts[-1]

        if k > 0:
            parent = tuple(parts[:-1])
            has_next = False
            for j in range(i+1, len(paths)):
                p2 = paths[j]
                if len(p2) < len(parent) or tuple(p2[:len(parent)]) != parent:
                    break
                if p2 != parts:
                    has_next = True
                    break
            vis[level_cols[k-1]] = "├──" if has_next else "└──"

        for a in range(0, max(0, k-1)):
            ancestor = tuple(parts[:a+1])
            vertical = False
            for j in range(i+1, len(paths)):
                p2 = paths[j]
                if len(p2) <= a or p2[a] != parts[a]:
                    break
                if tuple(p2[:a+1]) == ancestor:
                    vertical = True
                    break
            if vertical:
                vis[level_cols[a]] = "│"

        out.append({**vis, "Kind": r["Kind"], "MainRoot": r["MainRoot"], "Extra": r["Extra"]})
    return pd.DataFrame(out, columns=level_cols + ["Kind","MainRoot","Extra"])

# ===== 메인 =====
def main():
    main_roots = find_main_roots(root_dir)

    all_rows: List[Dict[str, Any]] = []
    all_summaries: List[pd.DataFrame] = []
    all_members: List[pd.DataFrame] = []

    for mr in main_roots:
        grouped, units = detect_groups(mr)
        collapsed = make_collapsed_rows(mr, grouped, units)
        for r in collapsed:
            r["MainRoot"] = os.path.basename(mr)
        all_rows.extend(collapsed)

        if not grouped.empty:
            all_summaries.append(grouped.assign(MainRoot=os.path.basename(mr)))

            mem_rows = []
            for _, g in grouped.iterrows():
                for pth, uname in zip(g["Parents"], g["UnitNames"]):
                    mem_rows.append({
                        "PatternID": g["PatternID"],
                        "MainRoot": g["MainRoot"],
                        "ParentPath": pth,
                        "UnitName": uname,
                        "FilesPerUnit": g["FilesPerUnit"],
                        "FileList": g["FileList"],
                        "ParentOfUnit": g["ParentOfUnit"],
                        "UnitsLabel": g["UnitsLabel"]
                    })
            all_members.append(pd.DataFrame(mem_rows))

    if not all_rows:
        print("No rows found.")
        return

    df = pd.DataFrame(all_rows)
    max_depth = max(len(p) for p in df["Parts"])
    for i in range(max_depth):
        df[f"Level{i}"] = df["Parts"].apply(lambda p, i=i: p[i] if i < len(p) else "")
    level_cols = [f"Level{i}" for i in range(max_depth)]
    df_paths = df[level_cols + ["Kind","MainRoot","Extra"]].copy()

    df_excel = build_box_matrix(df_paths, level_cols)

    # 요약/멤버
    df_summary = pd.concat(all_summaries, ignore_index=True) if all_summaries else pd.DataFrame(
        columns=["PatternID","MainRoot","ParentOfUnit","FileSet","Prefix","UnitCount","FilesPerUnit","UnitNames","Numbers","Parents","UnitsLabel","FileList"]
    )
    df_members = pd.concat(all_members, ignore_index=True) if all_members else pd.DataFrame(
        columns=["PatternID","MainRoot","ParentPath","UnitName","FilesPerUnit","FileList","ParentOfUnit","UnitsLabel"]
    )

    # ===== Excel 저장 =====
    with pd.ExcelWriter(excel_path, engine="openpyxl") as w:
        df_excel.to_excel(w, index=False, sheet_name="Tree_Collapsed")

        # (옵션) 한 컬럼 뷰
        tree_rows = []
        for _, r in df_paths.iterrows():
            parts = [r[c] for c in level_cols if r[c]]
            if not parts: continue
            if len(parts) == 1:
                tree_rows.append({"Tree": parts[0], "Kind": r["Kind"], "Extra": r["Extra"]})
            else:
                tree_rows.append({"Tree": ("    " * (len(parts)-1)) + "└── " + parts[-1],
                                  "Kind": r["Kind"], "Extra": r["Extra"]})
        pd.DataFrame(tree_rows, columns=["Tree","Kind","Extra"]).to_excel(w, index=False, sheet_name="Tree_View")

        if not df_summary.empty:
            df_summary[["PatternID","MainRoot","ParentOfUnit","UnitsLabel","UnitCount","FilesPerUnit","FileList"]].to_excel(
                w, index=False, sheet_name="PatternSummary")
        if not df_members.empty:
            df_members.to_excel(w, index=False, sheet_name="PatternMembers")

        # --- 스타일 ---
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment

        wb = w.book

        def style_ws(ws, wide_first=False):
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.font = Font(name=FONT_NAME, size=FONT_SIZE)
                    cell.alignment = Alignment(horizontal="left", vertical="center",
                                               wrap_text=False, shrink_to_fit=False)
            for c in range(1, ws.max_column + 1):
                letter = get_column_letter(c)
                max_len = 0
                for cell in ws[letter]:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                base = (LEVEL0_EXTRA_WIDTH if (c == 1 and wide_first) else DEFAULT_LEVEL_WIDTH)
                width = max(base, max_len + 1)
                width = min(EXCEL_MAX_COL_WIDTH, max(EXCEL_MIN_COL_WIDTH, width))
                ws.column_dimensions[letter].width = width
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        style_ws(wb["Tree_Collapsed"], wide_first=True)
        if "Tree_View" in wb.sheetnames: style_ws(wb["Tree_View"], wide_first=True)
        if "PatternSummary" in wb.sheetnames: style_ws(wb["PatternSummary"])
        if "PatternMembers" in wb.sheetnames: style_ws(wb["PatternMembers"])

    # ===== TXT (박스문자, .h 폴더 한 줄만) =====
    try:
        with open(txt_collapsed_path, "w", encoding="utf-8") as f:
            last_root = None
            for _, r in df_paths.iterrows():
                parts = [r[c] for c in level_cols if r[c]]
                if not parts: continue
                if parts[0] != last_root:
                    last_root = parts[0]
                    f.write("\n" + "="*80 + f"\n[{last_root}]\n" + "="*80 + "\n")
                    f.write(last_root + "\n")
                if len(parts) == 1: continue
                indent = "    " * (len(parts)-1)
                f.write(indent + "└── " + parts[-1] + "\n")
    except Exception as e:
        print("TXT write skipped:", e)

    print(f"✅ Excel saved: {excel_path}")
    print(f"✅ Collapsed text saved: {txt_collapsed_path}")

if __name__ == "__main__":
    main()
