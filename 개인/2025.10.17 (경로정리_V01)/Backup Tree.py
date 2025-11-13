# -*- coding: utf-8 -*-
"""
Backup 트리 (Collapsed + Depth Columns) - V18 (GUI)
- 진행률: '전체 step'을 작업 시작 전에 확정(PLAN) → done/total_step * 100
- 95% 캡/하드코딩 없음. SAVE 전에 100% 불가.
- SAVE 단계 상세 로그: 각 시트 작성/행 수/스타일/패딩/TXT 줄 수 등 세부 출력
- 기본 경로 기억/변경, 실시간 로그, 중지(부분 파일 정리), 프로세스 완전 종료
"""

import os, re, sys, math, json, queue, threading
from datetime import datetime
from typing import List, Dict, Any, Tuple, Set, Callable
import pandas as pd

# ===== 사용자 설정 =====
CASE_SENSITIVE = False
SQUASH_SPACES  = True
ROW_LIMIT      = None

IGNORE_DIR_KEYWORDS = {
    "__pycache__", ".git", ".venv", "node_modules",
    "$recycle.bin", "system volume information"
}
IGNORE_IMAGE_EXTS = {
    ".jpg",".jpeg",".png",".bmp",".gif",".tif",".tiff",
    ".webp",".heic",".heif",".raw",".cr2",".nef",".arw"
}
IGNORE_MISC_FILE_EXTS = {".tmp",".temp",".log~"}
IGNORE_FILE_EXTS = IGNORE_IMAGE_EXTS | IGNORE_MISC_FILE_EXTS

MAIN_ROOT_PATTERN = re.compile(r"^Backup_\d{4}\.\d{2}\.\d{2}(\b|[^\\/]*)", re.IGNORECASE)

# Excel 표시 파라미터(요청: 유지)
FONT_NAME = "Consolas"
FONT_SIZE = 11
EXCEL_MIN_COL_WIDTH = 10
EXCEL_MAX_COL_WIDTH = 12
DEFAULT_LEVEL_WIDTH = 14
LEVEL0_EXTRA_WIDTH = 12

def resource_path(rel_path: str) -> str:
    """PyInstaller/실행 스크립트 모두에서 리소스 경로 안전하게 가져오기"""
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, rel_path)

# ===== 설정 저장 =====
def _config_dir():
    base = os.getenv("APPDATA") or os.path.expanduser("~")
    p = os.path.join(base, "BackupTree"); os.makedirs(p, exist_ok=True); return p
def _config_path(): return os.path.join(_config_dir(), "config.json")
def load_config():
    try:
        with open(_config_path(), "r", encoding="utf-8") as f: return json.load(f)
    except Exception: return {}
def save_config(cfg):
    try:
        with open(_config_path(), "w", encoding="utf-8") as f: json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception: pass

# ===== 유틸 =====
def human_size(nbytes: int) -> str:
    if not nbytes: return "0 B"
    units = ["B","KB","MB","GB","TB","PB"]
    i = min(int(math.log(max(nbytes,1), 1024)), len(units)-1)
    return f"{round(nbytes/(1024**i), 2)} {units[i]}"

def safe_stat(path: str):
    try: return os.stat(path)
    except Exception: return None

def is_ignored_dir(name: str) -> bool:
    low = name.lower()
    return any(k in low for k in IGNORE_DIR_KEYWORDS)

def norm_ext(name: str) -> str:
    return os.path.splitext(name)[1].lower()

def normalize_filename(name: str) -> str:
    s = name.strip()
    if SQUASH_SPACES:
        s = re.sub(r"\s+", " ", s)
    if not CASE_SENSITIVE:
        s = s.lower()
    return s

def find_main_roots(root: str):
    roots = []
    try:
        for e in os.scandir(root):
            if e.is_dir(follow_symlinks=False) and MAIN_ROOT_PATTERN.match(e.name):
                roots.append(e.path)
    except Exception:
        pass
    roots.sort(key=lambda p: os.path.basename(p))
    return roots or [root]

def compress_numbers(nums: List[int]) -> str:
    if not nums: return ""
    nums = sorted(set(nums))
    parts = []
    start = prev = nums[0]
    for n in nums[1:]:
        if n == prev + 1:
            prev = n
        else:
            parts.append(str(start) if start==prev else f"{start}..{prev}")
            start = prev = n
    parts.append(str(start) if start==prev else f"{start}..{prev}")
    return ",".join(parts)

def split_prefix_number(name: str) -> Tuple[str, int]:
    m = re.match(r"^([^\d]*?)(\d+)$", name)
    if not m: return None, None
    return m.group(1), int(m.group(2))

def is_hiddle_folder(name: str) -> bool:
    return name.endswith(".h")

def hiddle_label(name: str) -> str:
    return (name[:-2] + " [Hiddle]") if is_hiddle_folder(name) else name

# ===== 진행률: 전체 step 고정 방식 =====
class Progress:
    """
    total_steps: 작업 시작 전에 PLAN 단계에서 확정
    done_steps : 각 tick 호출 시 누적
    퍼센트 = done_steps / total_steps * 100
    """
    def __init__(self, sink: Callable[[float, str], None]):
        self.sink = sink
        self.total_steps = 1
        self.done_steps = 0
        self._last_emit = -1
        # SAVE ETA(안내용)
        self._save_t_last = None
        self._save_ewma = None  # 초/틱

    def set_total_steps(self, total: int, note=""):
        self.total_steps = max(1, int(total))
        self._emit(f"[PLAN] 총 step = {self.total_steps:,} {note}".strip())

    def tick(self, msg=""):
        self.done_steps += 1
        self._emit(msg)

    def tick_save(self, msg=""):
        # ETA 추적(안내용)
        import time
        now = time.time()
        if self._save_t_last is not None:
            dt = max(1e-6, now - self._save_t_last)
            a = 0.3
            self._save_ewma = dt if self._save_ewma is None else (a*dt + (1-a)*self._save_ewma)
        self._save_t_last = now
        self.tick(msg)

    def eta_str(self, remain_ticks: int) -> str:
        if self._save_ewma is None: return "ETA --:--"
        secs = self._save_ewma * max(0, remain_ticks)
        m, s = divmod(int(secs), 60)
        return f"ETA {m:02d}:{s:02d}" if m < 100 else f"ETA {m:02d}m+"

    def _emit(self, msg=""):
        pct = (self.done_steps / self.total_steps) * 100.0
        pct = max(0.0, min(100.0, pct))
        if int(pct) != self._last_emit:
            self._last_emit = int(pct)
            self.sink(pct, msg)

# ===== 사전 카운트/드라이런(PLAN) =====
def prewalk_count(root: str) -> Tuple[int, int]:
    """ 방문 디렉터리 수(루트 포함), 유효 파일 수 """
    visited_dirs = 0
    file_cnt = 0
    def listdir_safe(p):
        try: return list(os.scandir(p))
        except Exception: return []
    stack = [root]
    while stack:
        d = stack.pop()
        visited_dirs += 1
        for e in listdir_safe(d):
            if e.is_dir(follow_symlinks=False):
                if is_ignored_dir(e.name): continue
                if not is_hiddle_folder(e.name):
                    stack.append(e.path)
            else:
                if norm_ext(e.name) in IGNORE_FILE_EXTS: continue
                file_cnt += 1
    return visited_dirs, file_cnt

def scan_tree_no_tick(root: str):
    """ 진행률 틱 없이 스캔(PLAN/DRY-RUN) """
    files_rows: List[Dict[str, Any]] = []
    units_rows: List[Dict[str, Any]] = []
    def listdir_sorted(path: str):
        try: return sorted(os.scandir(path), key=lambda e: (not e.is_dir(), e.name.lower()))
        except Exception: return []
    def build_structure_signature(direct_files_norm: Tuple[str, ...], child_dirs: List[str]) -> str:
        nonnum = []
        num_map: Dict[str, List[int]] = {}
        for dn in child_dirs:
            pre, num = split_prefix_number(dn)
            if pre is not None and num is not None: num_map.setdefault(pre, []).append(num)
            else: nonnum.append(dn)
        nonnum_sorted = tuple(sorted(nonnum))
        num_parts = [f"{pre}({compress_numbers(sorted(nums))})" for pre, nums in sorted(num_map.items())]
        files_sig = ",".join(direct_files_norm)
        return f"files=[{files_sig}]::dirs_nonnum=[{','.join(nonnum_sorted)}]::dirs_num=[{'|'.join(num_parts)}]"
    def dfs(path: str):
        base = os.path.basename(path)
        direct_files = []
        child_dirs = []
        for ch in listdir_sorted(path):
            if ch.is_dir(follow_symlinks=False):
                if is_ignored_dir(ch.name): continue
                child_dirs.append(ch.name)
            else:
                if norm_ext(ch.name) in IGNORE_FILE_EXTS: continue
                direct_files.append(normalize_filename(ch.name))
                st = safe_stat(ch.path)
                files_rows.append({
                    "Parent": path,
                    "Name": ch.name,
                    "NameNorm": normalize_filename(ch.name),
                    "Path": ch.path,
                    "Size(Bytes)": st.st_size if st else 0,
                    "Size(Human)": human_size(st.st_size) if st else ""
                })
        file_set = tuple(sorted(set(direct_files)))
        units_rows.append({
            "UnitPath": path,
            "ParentOfUnit": os.path.dirname(path),
            "UnitName": base,
            "FileSet": file_set,
            "FilesPerUnit": len(file_set),
            "NonNumChilds": tuple(sorted({n for n in child_dirs if split_prefix_number(n) == (None, None)})),
            "NumGroups": {}
        })
        pre_map: Dict[str, List[int]] = {}
        for dn in child_dirs:
            pre, num = split_prefix_number(dn)
            if pre is not None and num is not None: pre_map.setdefault(pre, []).append(num)
        units_rows[-1]["NumGroups"] = {k: sorted(v) for k, v in pre_map.items()}
        units_rows[-1]["StructureSig"] = build_structure_signature(file_set, child_dirs)
        if is_hiddle_folder(base): return
        for ch in listdir_sorted(path):
            if ch.is_dir(follow_symlinks=False):
                if is_ignored_dir(ch.name) or is_hiddle_folder(ch.name): continue
                dfs(ch.path)
    dfs(root)
    return files_rows, units_rows

def plan_build_rows(root: str, grouped: pd.DataFrame) -> int:
    """ BUILD 단계에서 추가될 row 수만 계산 (DRY) """
    count = 0
    parent_to_groups: Dict[str, List[dict]] = {}
    if grouped is not None and not grouped.empty:
        for _, g in grouped.iterrows():
            parent_to_groups.setdefault(g["ParentOfUnit"], []).append(g)
    grouped_member_paths: Set[str] = set()
    if grouped is not None and not grouped.empty:
        for parents in grouped["Parents"]:
            grouped_member_paths.update(parents)
    def listdir_sorted(path: str):
        try: return sorted(os.scandir(path), key=lambda e: (not e.is_dir(), e.name.lower()))
        except Exception: return []
    def dfs(path: str):
        nonlocal count
        base = os.path.basename(path)
        count += 1  # this folder row
        if is_hiddle_folder(base): return
        groups = parent_to_groups.get(path, [])
        if groups:
            for _ in groups:
                count += 1  # group line
        for ch in listdir_sorted(path):
            if ch.is_dir(follow_symlinks=False):
                if is_ignored_dir(ch.name): continue
                if ch.path in grouped_member_paths: continue
                if is_hiddle_folder(ch.name):
                    count += 1
                    continue
                dfs(ch.path)
            else:
                if norm_ext(ch.name) in IGNORE_FILE_EXTS: continue
                count += 1
    dfs(root)
    return count

def detect_groups_by_structure_no_tick(root: str, files_rows, units_rows):
    df_units = pd.DataFrame(units_rows)
    if df_units.empty:
        return pd.DataFrame(), df_units
    df_units["MainRoot"] = os.path.basename(root)
    df_units[["Prefix","Number"]] = df_units["UnitName"].apply(lambda n: pd.Series(split_prefix_number(n)))
    grouped = (
        df_units.groupby(["MainRoot","ParentOfUnit","StructureSig","Prefix"], dropna=False)
        .agg(
            UnitCount=("UnitName","nunique"),
            FilesPerUnit=("FilesPerUnit","first"),
            UnitNames=("UnitName", lambda x: sorted(set(x))),
            Numbers=("Number", lambda x: sorted(int(n) for n in x if pd.notna(n))),
            Parents=("UnitPath", lambda x: sorted(set(x)))
        )
        .reset_index()
    )
    def units_label(prefix, numbers, names):
        if pd.notna(prefix) and numbers:
            return f"{prefix}({compress_numbers(numbers)})"
        if len(names) <= 10:
            return ", ".join(names)
        return ", ".join(names[:10]) + f" … (+{len(names)-10})"
    grouped["UnitsLabel"] = grouped.apply(lambda r: units_label(r["Prefix"], r["Numbers"], r["UnitNames"]), axis=1)
    grouped["FileList"]   = grouped["StructureSig"].apply(lambda s: s.split("::")[0].replace("files=[","").replace("]",""))
    grouped = grouped[grouped["UnitCount"] >= 2].copy()
    grouped = grouped.sort_values(["ParentOfUnit","UnitCount"], ascending=[True, False])
    grouped["PatternID"] = (
        grouped.groupby("MainRoot").cumcount().add(1).astype(str)
        .radd(grouped["MainRoot"] + " :: P")
    )
    return grouped.reset_index(drop=True), df_units.reset_index(drop=True)

def plan_totals(root_dir: str, log=print) -> Dict[str, Any]:
    """
    PLAN(드라이런) 단계:
      - COUNT_total = 방문 디렉터리 수
      - SCAN_total  = (유효) 파일 수 + 방문 디렉터리 수
      - BUILD_total = Collapsed rows 수 (루트 포함)
      - SAVE_total  = df/행 수 기반 추정(시트/스타일/패딩/TXT 포함)
      - df_paths/summary/members는 SAVE_total 추정에 필요하므로 여기서 미리 계산
    """
    main_roots = find_main_roots(root_dir)
    total_count = 0
    total_scan  = 0
    total_build = 0

    all_rows: List[Dict[str, Any]] = []
    all_summaries: List[pd.DataFrame] = []
    all_members: List[pd.DataFrame] = []

    # COUNT/SCAN 총량
    for mr in main_roots:
        v_dirs, v_files = prewalk_count(mr)
        total_count += v_dirs
        total_scan  += (v_dirs + v_files)
        log(f"[PLAN] {os.path.basename(mr)}: dirs={v_dirs:,}, files={v_files:,}")

    # BUILD/SAVE 준비: 드라이런으로 rows/summary/members 산출
    def make_rows_for_plan(mr: str):
        files_rows, units_rows = scan_tree_no_tick(mr)
        grouped, _ = detect_groups_by_structure_no_tick(mr, files_rows, units_rows)
        # count collapsed rows
        build_rows = plan_build_rows(mr, grouped)
        rows = []

        # 실제 rows 구성(저장에 쓸 df_paths 만들려면 필요)
        parent_to_groups: Dict[str, List[dict]] = {}
        if not grouped.empty:
            for _, g in grouped.iterrows():
                parent_to_groups.setdefault(g["ParentOfUnit"], []).append(g)
        grouped_member_paths: Set[str] = set()
        if not grouped.empty:
            for parents in grouped["Parents"]:
                grouped_member_paths.update(parents)
        def listdir_sorted(path: str):
            try: return sorted(os.scandir(path), key=lambda e: (not e.is_dir(), e.name.lower()))
            except Exception: return []
        def dfs(path: str, parts: List[str]):
            base = os.path.basename(path)
            cur = parts[:]
            cur[-1] = hiddle_label(cur[-1])
            rows.append({"Parts": cur, "Kind":"Folder", "Name":cur[-1], "Extra":""})
            if is_hiddle_folder(base): return
            groups = parent_to_groups.get(path, [])
            if groups:
                for g in groups:
                    label = f"[Group] {g['UnitsLabel']}  [{g['UnitCount']} folders]"
                    rows.append({"Parts": cur[:] + [label], "Kind":"Group", "Name":g["UnitsLabel"], "Extra":""})
            for ch in listdir_sorted(path):
                if ch.is_dir(follow_symlinks=False):
                    if is_ignored_dir(ch.name): continue
                    if ch.path in grouped_member_paths: continue
                    if is_hiddle_folder(ch.name):
                        label = hiddle_label(ch.name)
                        rows.append({"Parts": cur[:] + [label], "Kind":"Folder", "Name":label, "Extra":""})
                        continue
                    dfs(ch.path, cur+[ch.name])
                else:
                    if norm_ext(ch.name) in IGNORE_FILE_EXTS: continue
                    st = safe_stat(ch.path)
                    rows.append({"Parts": cur[:] + [ch.name], "Kind":"File", "Name":ch.name,
                                "Extra":human_size(st.st_size) if st else ""})
        dfs(mr, [os.path.basename(mr)])
        # members table
        mem_rows = []
        if not grouped.empty:
            for _, g in grouped.iterrows():
                for uname, pth in zip(g["UnitNames"], g["Parents"]):
                    mem_rows.append({
                        "PatternID": g["PatternID"],
                        "MainRoot": g["MainRoot"],
                        "ParentPath": pth,
                        "UnitName": uname,
                        "FilesPerUnit": g["FilesPerUnit"],
                        "UnitsLabel": g["UnitsLabel"],
                        "ParentOfUnit": g["ParentOfUnit"]
                    })
        return rows, grouped.assign(MainRoot=os.path.basename(mr)), pd.DataFrame(mem_rows), build_rows

    for mr in main_roots:
        rows, grouped, members, build_rows = make_rows_for_plan(mr)
        total_build += build_rows
        for r in rows: r["MainRoot"] = os.path.basename(mr)
        all_rows.extend(rows)
        if not grouped.empty: all_summaries.append(grouped)
        if not members.empty: all_members.append(members)

    # df_paths/level_cols for save estimation
    df = pd.DataFrame(all_rows) if all_rows else pd.DataFrame(columns=["Parts","Kind","Name","Extra","MainRoot"])
    max_depth = max((len(p) for p in df["Parts"]), default=1)
    for i in range(max_depth):
        df[f"Level{i}"] = df["Parts"].apply(lambda p, i=i: (p[i] if i < len(p) else ""))
    level_cols = [f"Level{i}" for i in range(max_depth)]
    df_paths = df[level_cols + ["Kind","MainRoot","Extra"]].copy()

    df_summary = pd.concat(all_summaries, ignore_index=True) if all_summaries else pd.DataFrame()
    df_members = pd.concat(all_members,   ignore_index=True) if all_members   else pd.DataFrame()

    # SAVE_total 추정
    save_total = estimate_save_total(df_paths, df_summary, df_members)

    total_steps = total_count + total_scan + total_build + save_total
    plan = {
        "total_steps": total_steps,
        "count_total": total_count,
        "scan_total": total_scan,
        "build_total": total_build,
        "save_total": save_total,
        "df_paths": df_paths,
        "level_cols": level_cols,
        "df_summary": df_summary,
        "df_members": df_members,
        "main_roots": main_roots
    }
    return plan

# ===== 실제 스캔/빌드(틱 발생) =====
def scan_tree(root: str, check_stop: Callable[[], bool], progress: Progress, log=print):
    files_rows: List[Dict[str, Any]] = []
    units_rows: List[Dict[str, Any]] = []
    def listdir_sorted(path: str):
        try: return sorted(os.scandir(path), key=lambda e: (not e.is_dir(), e.name.lower()))
        except Exception: return []
    def build_structure_signature(direct_files_norm: Tuple[str, ...], child_dirs: List[str]) -> str:
        nonnum = []
        num_map: Dict[str, List[int]] = {}
        for dn in child_dirs:
            pre, num = split_prefix_number(dn)
            if pre is not None and num is not None:
                num_map.setdefault(pre, []).append(num)
            else:
                nonnum.append(dn)
        nonnum_sorted = tuple(sorted(nonnum))
        num_parts = [f"{pre}({compress_numbers(sorted(nums))})" for pre, nums in sorted(num_map.items())]
        files_sig = ",".join(direct_files_norm)
        return f"files=[{files_sig}]::dirs_nonnum=[{','.join(nonnum_sorted)}]::dirs_num=[{'|'.join(num_parts)}]"
    def dfs(path: str):
        if check_stop(): raise RuntimeError("사용자 중지")
        base = os.path.basename(path)
        direct_files = []
        child_dirs = []
        for ch in listdir_sorted(path):
            if ch.is_dir(follow_symlinks=False):
                if is_ignored_dir(ch.name): continue
                child_dirs.append(ch.name)
            else:
                if norm_ext(ch.name) in IGNORE_FILE_EXTS: continue
                direct_files.append(normalize_filename(ch.name))
                st = safe_stat(ch.path)
                files_rows.append({
                    "Parent": path, "Name": ch.name, "NameNorm": normalize_filename(ch.name),
                    "Path": ch.path, "Size(Bytes)": st.st_size if st else 0,
                    "Size(Human)": human_size(st.st_size) if st else ""
                })
                progress.tick(f"[SCAN][FILE] {ch.path}")
        progress.tick(f"[SCAN][DIR ] {path}")  # 방문 tick
        file_set = tuple(sorted(set(direct_files)))
        units_rows.append({
            "UnitPath": path, "ParentOfUnit": os.path.dirname(path), "UnitName": base,
            "FileSet": file_set, "FilesPerUnit": len(file_set), "NonNumChilds": tuple(sorted({n for n in child_dirs if split_prefix_number(n) == (None, None)})),
            "NumGroups": {}
        })
        pre_map: Dict[str, List[int]] = {}
        for dn in child_dirs:
            pre, num = split_prefix_number(dn)
            if pre is not None and num is not None: pre_map.setdefault(pre, []).append(num)
        units_rows[-1]["NumGroups"] = {k: sorted(v) for k, v in pre_map.items()}
        units_rows[-1]["StructureSig"] = build_structure_signature(file_set, child_dirs)
        if is_hiddle_folder(base): return
        for ch in listdir_sorted(path):
            if ch.is_dir(follow_symlinks=False):
                if is_ignored_dir(ch.name) or is_hiddle_folder(ch.name): continue
                dfs(ch.path)
    dfs(root)
    return files_rows, units_rows

def detect_groups_by_structure(root: str, files_rows, units_rows):
    df_units = pd.DataFrame(units_rows)
    if df_units.empty: return pd.DataFrame(), df_units
    df_units["MainRoot"] = os.path.basename(root)
    df_units[["Prefix","Number"]] = df_units["UnitName"].apply(lambda n: pd.Series(split_prefix_number(n)))
    grouped = (
        df_units.groupby(["MainRoot","ParentOfUnit","StructureSig","Prefix"], dropna=False)
        .agg(
            UnitCount=("UnitName","nunique"),
            FilesPerUnit=("FilesPerUnit","first"),
            UnitNames=("UnitName", lambda x: sorted(set(x))),
            Numbers=("Number", lambda x: sorted(int(n) for n in x if pd.notna(n))),
            Parents=("UnitPath", lambda x: sorted(set(x)))
        )
        .reset_index()
    )
    def units_label(prefix, numbers, names):
        if pd.notna(prefix) and numbers: return f"{prefix}({compress_numbers(numbers)})"
        if len(names) <= 10: return ", ".join(names)
        return ", ".join(names[:10]) + f" … (+{len(names)-10})"
    grouped["UnitsLabel"] = grouped.apply(lambda r: units_label(r["Prefix"], r["Numbers"], r["UnitNames"]), axis=1)
    grouped["FileList"]   = grouped["StructureSig"].apply(lambda s: s.split("::")[0].replace("files=[","").replace("]",""))
    grouped = grouped[grouped["UnitCount"] >= 2].copy()
    grouped = grouped.sort_values(["ParentOfUnit","UnitCount"], ascending=[True, False])
    grouped["PatternID"] = (
        grouped.groupby("MainRoot").cumcount().add(1).astype(str)
        .radd(grouped["MainRoot"] + " :: P")
    )
    return grouped.reset_index(drop=True), df_units.reset_index(drop=True)

def make_collapsed_rows(root: str, grouped: pd.DataFrame, check_stop: Callable[[], bool],
                        progress: Progress, log=print) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if grouped is None or not isinstance(grouped, pd.DataFrame) or grouped.empty:
        grouped = pd.DataFrame()
    parent_to_groups: Dict[str, List[dict]] = {}
    if not grouped.empty:
        for _, g in grouped.iterrows():
            parent_to_groups.setdefault(g["ParentOfUnit"], []).append(g)
    grouped_member_paths: Set[str] = set()
    if not grouped.empty:
        for parents in grouped["Parents"]:
            grouped_member_paths.update(parents)
    def listdir_sorted(path: str):
        try: return sorted(os.scandir(path), key=lambda e: (not e.is_dir(), e.name.lower()))
        except Exception: return []
    def dfs(path: str, parts: List[str]):
        if check_stop(): raise RuntimeError("사용자 중지")
        base = os.path.basename(path)
        cur = parts[:]; cur[-1] = hiddle_label(cur[-1])
        rows.append({"Parts": cur, "Kind":"Folder", "Name":cur[-1], "Extra":""})
        progress.tick(f"[BUILD][ROW ] {path}")
        if ROW_LIMIT and len(rows) >= ROW_LIMIT: return
        if is_hiddle_folder(base): return
        groups = parent_to_groups.get(path, [])
        if groups:
            for g in groups:
                label = f"[Group] {g['UnitsLabel']}  [{g['UnitCount']} folders]"
                rows.append({"Parts": cur[:] + [label], "Kind":"Group", "Name":g["UnitsLabel"], "Extra":""})
                progress.tick(f"[BUILD][GRUP] {label}")
                if ROW_LIMIT and len(rows) >= ROW_LIMIT: return
        for ch in listdir_sorted(path):
            if ch.is_dir(follow_symlinks=False):
                if is_ignored_dir(ch.name): continue
                if ch.path in grouped_member_paths: continue
                if is_hiddle_folder(ch.name):
                    label = hiddle_label(ch.name)
                    rows.append({"Parts": cur[:] + [label], "Kind":"Folder", "Name":label, "Extra":""})
                    progress.tick(f"[BUILD][HIDE] {label}")
                    if ROW_LIMIT and len(rows) >= ROW_LIMIT: return
                    continue
                dfs(ch.path, cur+[ch.name])
                if ROW_LIMIT and len(rows) >= ROW_LIMIT: return
            else:
                if norm_ext(ch.name) in IGNORE_FILE_EXTS: continue
                st = safe_stat(ch.path)
                rows.append({"Parts": cur[:] + [ch.name], "Kind":"File", "Name":ch.name,
                            "Extra":human_size(st.st_size) if st else ""})
                progress.tick(f"[BUILD][FILE] {ch.name}")
                if ROW_LIMIT and len(rows) >= ROW_LIMIT: return
    dfs(root, [os.path.basename(root)])
    return rows

# ===== 박스문자/패딩 =====
def build_box_matrix(df_paths: pd.DataFrame, level_cols: List[str]) -> pd.DataFrame:
    paths = [[r[c] for c in level_cols if r[c]] for _, r in df_paths.iterrows()]
    out = []
    for i, r in df_paths.iterrows():
        parts = paths[i]
        vis = {c: "" for c in level_cols}
        if not parts:
            out.append({**vis, "Kind": r["Kind"], "MainRoot": r["MainRoot"], "Extra": r["Extra"]})
            continue
        k = len(parts) - 1
        vis[level_cols[k]] = parts[-1]
        if k > 0:
            parent = tuple(parts[:-1]); has_next = False
            for j in range(i+1, len(paths)):
                p2 = paths[j]
                if len(p2) < len(parent) or tuple(p2[:len(parent)]) != parent: break
                if p2 != parts: has_next = True; break
            vis[level_cols[k-1]] = "├" if has_next else "└"
        for a in range(0, max(0, k-1)):
            ancestor = tuple(parts[:a+1]); vertical = False
            for j in range(i+1, len(paths)):
                p2 = paths[j]
                if len(p2) <= a or p2[a] != parts[a]: break
                if tuple(p2[:a+1]) == ancestor: vertical = True; break
            if vertical: vis[level_cols[a]] = "│"
        out.append({**vis, "Kind": r["Kind"], "MainRoot": r["MainRoot"], "Extra": r["Extra"]})
    return pd.DataFrame(out, columns=level_cols + ["Kind","MainRoot","Extra"])

def pad_connectors_to_width(ws, level_col_count: int, margin_chars: int = 1):
    from openpyxl.utils import get_column_letter
    target_len = {}
    for c in range(1, level_col_count + 1):
        letter = get_column_letter(c)
        width = ws.column_dimensions[letter].width or 10
        target_len[c] = max(1, int(round(width)) - margin_chars)
    for row in range(2, ws.max_row + 1):
        for c in range(1, level_col_count):
            cell = ws.cell(row=row, column=c)
            if cell.value is None:
                continue
            s = str(cell.value)
            if not s:
                continue
            if s[0] in ("├", "└"):
                head = s[0]
                cell.value = head + ("─" * max(0, target_len[c] - 1))

# ===== SAVE 총량 추정 =====
def estimate_save_total(df_paths: pd.DataFrame, df_summary: pd.DataFrame, df_members: pd.DataFrame) -> int:
    total = 0
    # 시트 작성: 4개(있을 때만 카운트)
    total += 1  # Tree_Collapsed 쓰기
    total += 1  # Tree_View 쓰기
    total += 1  # PatternSummary
    total += 1  # PatternMembers
    # 구성/배치 중간 tick(행 기반)
    total += max(1, len(df_paths) // 500)    # Tree_View 구성 중간 보고
    # 스타일 4
    total += 4
    # 커넥터 패딩 1
    total += 1
    # TXT 저장 + 중간 tick
    total += 1 + max(1, len(df_paths) // 2000)
    return total

# ===== 저장(세부 로그) =====
def save_outputs(root_dir: str, df_paths: pd.DataFrame, level_cols: List[str],
                 df_summary: pd.DataFrame, df_members: pd.DataFrame,
                 excel_basename: str,
                 check_stop: Callable[[], bool], progress: Progress, log=print):
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment

    excel_path = os.path.join(root_dir, excel_basename + ".xlsx")
    txt_collapsed_path = os.path.join(root_dir, excel_basename + "_collapsed.txt")

    # === Excel 저장 ===
    import openpyxl
    with pd.ExcelWriter(excel_path, engine="openpyxl") as w:
        # 1) Tree_Collapsed
        df_excel = build_box_matrix(df_paths, level_cols)
        df_excel.to_excel(w, index=False, sheet_name="Tree_Collapsed")
        progress.tick_save(f"[SAVE] Tree_Collapsed 시트 작성 완료 ({len(df_excel):,} 행)")

        # 2) Tree_View (중간 진행 로그 포함)
        tree_rows = []
        for idx, r in df_paths.iterrows():
            if check_stop(): raise RuntimeError("사용자 중지")
            parts = [r[c] for c in level_cols if r[c]]
            if not parts:
                if (idx+1) % 500 == 0:
                    remain = 0  # 구성은 다음 tick에서 반영됨
                    log(f"[SAVE] Tree_View 구성 진행: {idx+1:,} 행")
                    progress.tick_save(f"[SAVE] Tree_View 구성 진행 {idx+1:,}행 {progress.eta_str(remain)}")
                continue
            if len(parts) == 1:
                tree_rows.append({"Tree": parts[0], "Kind": r["Kind"], "Extra": r["Extra"]})
            else:
                tree_rows.append({"Tree": ("    " * (len(parts)-1)) + "└── " + parts[-1],
                                  "Kind": r["Kind"], "Extra": r["Extra"]})
            if (idx+1) % 500 == 0:
                log(f"[SAVE] Tree_View 구성 진행: {idx+1:,} 행")
                progress.tick_save(f"[SAVE] Tree_View 구성 진행 {idx+1:,}행")
        df_tree = pd.DataFrame(tree_rows, columns=["Tree","Kind","Extra"])
        df_tree.to_excel(w, index=False, sheet_name="Tree_View")
        progress.tick_save(f"[SAVE] Tree_View 시트 작성 완료 ({len(df_tree):,} 행)")

        # 3) PatternSummary
        if not df_summary.empty:
            cols = ["PatternID","MainRoot","ParentOfUnit","UnitsLabel","UnitCount"]
            use_cols = [c for c in cols if c in df_summary.columns]
            df_summary[use_cols].to_excel(w, index=False, sheet_name="PatternSummary")
            progress.tick_save(f"[SAVE] PatternSummary 시트 작성 완료 ({len(df_summary):,} 행)")
        else:
            progress.tick_save(f"[SAVE] PatternSummary 비어있음 (건너뜀)")

        # 4) PatternMembers
        if not df_members.empty:
            df_members.to_excel(w, index=False, sheet_name="PatternMembers")
            progress.tick_save(f"[SAVE] PatternMembers 시트 작성 완료 ({len(df_members):,} 행)")
        else:
            progress.tick_save(f"[SAVE] PatternMembers 비어있음 (건너뜀)")

        # 5) 스타일
        wb = w.book
        def style_ws(ws, wide_first=False, name=""):
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.font = Font(name=FONT_NAME, size=FONT_SIZE)
                    cell.alignment = Alignment(horizontal="left", vertical="center",
                                               wrap_text=False, shrink_to_fit=False)
            for c in range(1, ws.max_column + 1):
                letter = get_column_letter(c)
                max_len = 0
                for cell in ws[letter]:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                base = (LEVEL0_EXTRA_WIDTH if (c == 1 and wide_first) else DEFAULT_LEVEL_WIDTH)
                width = max(base, max_len + 1)
                width = min(EXCEL_MAX_COL_WIDTH, max(EXCEL_MIN_COL_WIDTH, width))
                ws.column_dimensions[letter].width = width
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            progress.tick_save(f"[SAVE] 스타일 적용: {name} (rows={ws.max_row:,}, cols={ws.max_column})")

        if "Tree_Collapsed" in wb.sheetnames: style_ws(wb["Tree_Collapsed"], True,  "Tree_Collapsed")
        if "Tree_View"      in wb.sheetnames: style_ws(wb["Tree_View"],      True,  "Tree_View")
        if "PatternSummary" in wb.sheetnames: style_ws(wb["PatternSummary"], False, "PatternSummary")
        if "PatternMembers" in wb.sheetnames: style_ws(wb["PatternMembers"], False, "PatternMembers")

        # 6) 커넥터 패딩
        if "Tree_Collapsed" in wb.sheetnames:
            ws_tc = wb["Tree_Collapsed"]
            level_col_count = sum(1 for c in ws_tc[1] if str(c.value).startswith("Level"))
            pad_connectors_to_width(ws_tc, level_col_count, margin_chars=1)
            progress.tick_save(f"[SAVE] 커넥터 패딩 적용 (level_cols={level_col_count})")

    progress.tick_save("[SAVE] 엑셀 파일 저장 완료")

    # === TXT 저장 ===
    with open(txt_collapsed_path, "w", encoding="utf-8") as f:
        last_root = None
        line_count = 0
        for idx, r in df_paths.iterrows():
            if check_stop(): raise RuntimeError("사용자 중지")
            part_cols = [c for c in df_paths.columns if c.startswith("Level")]
            parts = [r[c] for c in part_cols if r[c]]
            if not parts:
                if (idx+1) % 2000 == 0:
                    progress.tick_save(f"[SAVE] TXT 진행 {idx+1:,}행")
                continue
            if parts[0] != last_root:
                last_root = parts[0]
                f.write("\n" + "="*80 + f"\n[{last_root}]\n" + "="*80 + "\n")
                f.write(last_root + "\n"); line_count += 3
            if len(parts) > 1:
                indent = "    " * (len(parts)-1)
                f.write(indent + "└── " + parts[-1] + "\n")
                line_count += 1
            if (idx+1) % 2000 == 0:
                progress.tick_save(f"[SAVE] TXT 진행 {idx+1:,}행")
        progress.tick_save(f"[SAVE] TXT 저장 완료 (총 {line_count:,} 줄)")

    return excel_path, txt_collapsed_path

# ===== 메인 파이프라인 =====
def run_backup_tree(root_dir: str, check_stop: Callable[[], bool],
                    emit_progress: Callable[[float, str], None], log=print):
    if not root_dir or not os.path.isdir(root_dir):
        raise ValueError("유효한 root_dir 폴더를 선택/입력하세요.")
    log(f"[INIT] root_dir = {root_dir}")
    emit_progress(0.0, "초기화")

    # 0) PLAN: 전체 step 확정 + df 준비
    plan = plan_totals(root_dir, log=log)
    total_steps = plan["total_steps"]
    df_paths = plan["df_paths"]; level_cols = plan["level_cols"]
    df_summary = plan["df_summary"]; df_members = plan["df_members"]
    main_roots = plan["main_roots"]

    progress = Progress(emit_progress)
    progress.set_total_steps(total_steps, note=f"(COUNT={plan['count_total']:,}, SCAN={plan['scan_total']:,}, BUILD={plan['build_total']:,}, SAVE={plan['save_total']:,})")

    # 1) 실제 SCAN/BUILD 실행(틱 발생) ─ df_paths는 PLAN에서 이미 구성되어 있으므로 BUILD 단계는 '행 생성 과정'에 맞추어 다시 틱
    all_rows: List[Dict[str, Any]] = []
    all_summaries: List[pd.DataFrame] = []
    all_members: List[pd.DataFrame] = []

    for mr in main_roots:
        if check_stop(): raise RuntimeError("사용자 중지")
        log(f"[SCAN] {mr}")
        files_rows, units_rows = scan_tree(mr, check_stop, progress, log)
        grouped, _ = detect_groups_by_structure(mr, files_rows, units_rows)
        rows = make_collapsed_rows(mr, grouped, check_stop, progress, log)
        for r in rows: r["MainRoot"] = os.path.basename(mr)
        all_rows.extend(rows)
        if not grouped.empty:
            all_summaries.append(grouped.assign(MainRoot=os.path.basename(mr)))
            mem_rows = []
            for _, g in grouped.iterrows():
                for uname, pth in zip(g["UnitNames"], g["Parents"]):
                    mem_rows.append({
                        "PatternID": g["PatternID"], "MainRoot": g["MainRoot"],
                        "ParentPath": pth, "UnitName": uname, "FilesPerUnit": g["FilesPerUnit"],
                        "UnitsLabel": g["UnitsLabel"], "ParentOfUnit": g["ParentOfUnit"]
                    })
            all_members.append(pd.DataFrame(mem_rows))

    if check_stop(): raise RuntimeError("사용자 중지")
    if not all_rows:
        raise RuntimeError("스캔 결과가 없습니다.")

    # 저장 (df_paths/summary/members는 PLAN 결과 사용 → SAVE 분모는 이미 total에 포함됨)
    save_basename = f"_BackupTree_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    log("[SAVE] 파일 저장 시작")
    excel_path, txt_path = save_outputs(root_dir, df_paths, level_cols,
                                        df_summary, df_members, save_basename,
                                        check_stop, progress, log)

    emit_progress(100.0, "완료")
    log("[DONE] 저장 완료")
    log(f"[DONE] Excel: {excel_path}")
    log(f"[DONE] TXT  : {txt_path}")
    return excel_path, txt_path

# ===== GUI =====
def launch_gui():
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox

    class App:
        def __init__(self, master):
            self.master = master
            master.title("Backup Tree")
            master.geometry("900x600")

            self.q = queue.Queue()
            self.worker = None
            self._stop_flag = False

            cfg = load_config()
            default_root = cfg.get("default_root") or os.path.expanduser("~")

            # 경로 영역
            top = ttk.Frame(master); top.pack(fill="x", padx=10, pady=10)
            ttk.Label(top, text="루트 경로:").grid(row=0, column=0, sticky="w")
            self.var_root = tk.StringVar(value=default_root)
            self.ent_root = ttk.Entry(top, textvariable=self.var_root, width=90)
            self.ent_root.grid(row=0, column=1, sticky="we", padx=(6,6))
            top.columnconfigure(1, weight=1)
            self.btn_change = ttk.Button(top, text="경로 변경…", command=self.on_change_root)
            self.btn_change.grid(row=0, column=2, sticky="e")
            self.var_save_default = tk.BooleanVar(value=False)
            self.chk_default = ttk.Checkbutton(top, text="이 경로를 기본값으로 저장", variable=self.var_save_default)
            self.chk_default.grid(row=1, column=1, sticky="w", pady=(6,0))

            # 실행/중지
            mid = ttk.Frame(master); mid.pack(fill="x", padx=10, pady=(0,6))
            self.btn_run = ttk.Button(mid, text="실행", command=self.on_run); self.btn_run.pack(side="left")
            self.btn_stop = ttk.Button(mid, text="중지", command=self.on_stop, state="disabled"); self.btn_stop.pack(side="left", padx=(8,0))

            # 진행률
            pr = ttk.Frame(master); pr.pack(fill="x", padx=10, pady=(0,6))
            self.var_pct = tk.StringVar(value="0%")
            ttk.Label(pr, text="진행률:").pack(side="left")
            self.lbl_pct = ttk.Label(pr, textvariable=self.var_pct, width=6); self.lbl_pct.pack(side="left", padx=(6,12))
            self.pb = ttk.Progressbar(pr, mode="determinate", maximum=100); self.pb.pack(fill="x", expand=True)

            # 로그
            frm_log = ttk.LabelFrame(master, text="실시간 로그"); frm_log.pack(fill="both", expand=True, padx=10, pady=(0,10))
            self.txt = tk.Text(frm_log, height=26, wrap="word")
            self.txt.pack(fill="both", expand=True)

            master.protocol("WM_DELETE_WINDOW", self.on_close)
            self.master.after(75, self.poll_queue)

            self.log("[READY] 경로 확인 후 [실행]을 누르세요.")

        def log(self, s: str):
            self.q.put(("log", s))

        def set_progress(self, pct: float, msg: str):
            self.q.put(("progress", (pct, msg)))

        def poll_queue(self):
            try:
                while True:
                    kind, payload = self.q.get_nowait()
                    if kind == "log":
                        self.txt.insert("end", payload + "\n"); self.txt.see("end")
                    elif kind == "progress":
                        pct, msg = payload
                        self.pb['value'] = pct
                        self.var_pct.set(f"{int(pct)}%")
                        if msg:
                            self.txt.insert("end", f"{int(pct)}% - {msg}\n"); self.txt.see("end")
                    elif kind == "done":
                        self.btn_run.config(state="normal"); self.btn_stop.config(state="disabled")
                        excel, txtp = payload
                        self.pb['value'] = 100; self.var_pct.set("100%")
                        messagebox.showinfo("완료", f"작업 완료!\n\nExcel:\n{excel}\n\nTXT:\n{txtp}")
                    elif kind == "error":
                        self.btn_run.config(state="normal"); self.btn_stop.config(state="disabled")
                        messagebox.showerror("오류", str(payload))
                    self.q.task_done()
            except queue.Empty:
                pass
            self.master.after(50, self.poll_queue)

        def on_change_root(self):
            d = filedialog.askdirectory(title="루트 경로 선택", initialdir=self.var_root.get() or os.path.expanduser("~"))
            if d: self.var_root.set(d)

        def on_run(self):
            root_dir = self.var_root.get().strip()
            if not root_dir or not os.path.isdir(root_dir):
                from tkinter import messagebox
                messagebox.showerror("유효하지 않은 경로", "존재하는 폴더를 선택하세요.")
                return
            if self.var_save_default.get():
                cfg = load_config(); cfg["default_root"] = root_dir; save_config(cfg)
                self.log(f"[CONFIG] 기본 경로 저장: {root_dir}")

            self._stop_flag = False
            self.btn_run.config(state="disabled"); self.btn_stop.config(state="normal")
            self.pb['value'] = 0; self.var_pct.set("0%")
            self.txt.insert("end", "\n=== 작업 시작 ===\n")

            def job():
                try:
                    def check_stop(): return self._stop_flag
                    def emit_progress(pct, msg): self.set_progress(pct, msg)
                    def gui_log(*args): self.log(" ".join(str(a) for a in args))
                    excel_path, txt_path = run_backup_tree(root_dir, check_stop, emit_progress, log=gui_log)
                    self.q.put(("done", (excel_path, txt_path)))
                except Exception as e:
                    self.q.put(("error", e))
            self.worker = threading.Thread(target=job, daemon=True); self.worker.start()

        def on_stop(self):
            self._stop_flag = True
            self.btn_stop.config(state="disabled")
            self.log("[STOP] 중지 요청 → 가능한 지점에서 즉시 중단 및 부분 파일 정리")

        def on_close(self):
            self._stop_flag = True
            if self.worker and self.worker.is_alive():
                try: self.worker.join(timeout=2.0)
                except Exception: pass
            try:
                self.master.destroy()
            finally:
                os._exit(0)

    # DPI 보정(Windows)
    try:
        if sys.platform.startswith("win"):
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass

    import tkinter as tk
    root = tk.Tk()
    
    # --- Windows 작업표시줄 아이콘/그룹 고정 (AppUserModelID) ---
    try:
        if sys.platform.startswith("win"):
            import ctypes
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(
                "com.backuptree.app"  # 임의의 고유 문자열
            )
    except Exception as e:
        print(f"[ICON] AppUserModelID 설정 실패: {e}")

    # --- 아이콘 경로 ---
    ico_path = resource_path("app.ico")
    png_path = resource_path("ico.png")

    # --- 작업표시줄/제목표시줄: .ico (Windows 전용) ---
    try:
        if sys.platform.startswith("win") and os.path.exists(ico_path):
            root.iconbitmap(default=ico_path)  # 작업 표시줄은 EXE/ICO를 참조
    except Exception as e:
        print(f"[ICON] iconbitmap 실패: {e}")

    # --- 크로스 플랫폼/Alt-Tab: PNG ---
    try:
        if os.path.exists(png_path):
            import tkinter as _tk
            _icon_img = _tk.PhotoImage(file=png_path)
            root.iconphoto(True, _icon_img)
    except Exception as e:
        print(f"[ICON] iconphoto 실패: {e}")


    App(root)
    root.mainloop()

# ===== 실행 진입점 =====
if __name__ == "__main__":
    if len(sys.argv) > 1:
        target = sys.argv[1]
        def emit(p, m): print(f"{int(p):3d}% - {m}")
        def stop(): return False
        def lg(*a): print(" ".join(str(x) for x in a))
        run_backup_tree(target, stop, emit, log=lg)
    else:
        launch_gui()
