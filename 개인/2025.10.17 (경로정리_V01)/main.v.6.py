# -*- coding: utf-8 -*-
"""
Backup 트리 (Collapsed + Depth Columns) - V17d
- 요구 반영:
  * SAVE 단계 시작 즉시 95% 탈출(시작 틱)
  * SAVE 단계 세부 진행률(자주 tick) + ETA(남은 시간) 표시
  * [중지] 즉시 반영(부분 파일 정리 후 중단)
  * 기본 경로 저장/불러오기, 실시간 로그, 프로세스 완전 종료 유지
"""

import os
import re
import sys
import math
import json
import queue
import threading
from datetime import datetime
from typing import List, Dict, Any, Tuple, Set, Callable

import pandas as pd

# ======================= 사용자 기본 설정 (로직 변경 없음) =======================
CASE_SENSITIVE = False
SQUASH_SPACES  = True
ROW_LIMIT      = None

IGNORE_DIR_KEYWORDS = {
    "__pycache__", ".git", ".venv", "node_modules",
    "$recycle.bin", "system volume information"
}
IGNORE_IMAGE_EXTS = {
    ".jpg",".jpeg",".png",".bmp",".gif",".tif",".tiff",
    ".webp",".heic",".heif",".raw",".cr2",".nef",".arw"
}
IGNORE_MISC_FILE_EXTS = {".tmp",".temp",".log~"}
IGNORE_FILE_EXTS = IGNORE_IMAGE_EXTS | IGNORE_MISC_FILE_EXTS

MAIN_ROOT_PATTERN = re.compile(r"^Backup_\d{4}\.\d{2}\.\d{2}(\b|[^\\/]*)", re.IGNORECASE)

# ======= Excel 표시 파라미터 (요청: 변경하지 않음) =======
FONT_NAME = "Consolas"
FONT_SIZE = 11
EXCEL_MIN_COL_WIDTH = 10
EXCEL_MAX_COL_WIDTH = 12
DEFAULT_LEVEL_WIDTH = 14
LEVEL0_EXTRA_WIDTH = 12
# =============================================================================

# ======================= 환경/설정 저장 =======================
def _config_dir() -> str:
    base = os.getenv("APPDATA") or os.path.expanduser("~")
    path = os.path.join(base, "BackupTree")
    os.makedirs(path, exist_ok=True)
    return path

def _config_path() -> str:
    return os.path.join(_config_dir(), "config.json")

def load_config() -> Dict[str, Any]:
    try:
        with open(_config_path(), "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def save_config(cfg: Dict[str, Any]):
    try:
        with open(_config_path(), "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

# ======================= 유틸 =======================
def human_size(nbytes: int) -> str:
    if not nbytes: return "0 B"
    units = ["B","KB","MB","GB","TB","PB"]
    i = min(int(math.log(max(nbytes,1), 1024)), len(units)-1)
    return f"{round(nbytes/(1024**i), 2)} {units[i]}"

def safe_stat(path: str):
    try: return os.stat(path)
    except Exception: return None

def is_ignored_dir(name: str) -> bool:
    low = name.lower()
    return any(k in low for k in IGNORE_DIR_KEYWORDS)

def norm_ext(name: str) -> str:
    return os.path.splitext(name)[1].lower()

def normalize_filename(name: str) -> str:
    s = name.strip()
    if SQUASH_SPACES:
        s = re.sub(r"\s+", " ", s)
    if not CASE_SENSITIVE:
        s = s.lower()
    return s

def find_main_roots(root: str):
    roots = []
    try:
        for e in os.scandir(root):
            if e.is_dir(follow_symlinks=False) and MAIN_ROOT_PATTERN.match(e.name):
                roots.append(e.path)
    except Exception:
        pass
    roots.sort(key=lambda p: os.path.basename(p))
    return roots or [root]

def compress_numbers(nums: List[int]) -> str:
    if not nums: return ""
    nums = sorted(set(nums))
    parts = []
    start = prev = nums[0]
    for n in nums[1:]:
        if n == prev + 1:
            prev = n
        else:
            parts.append(str(start) if start==prev else f"{start}..{prev}")
            start = prev = n
    parts.append(str(start) if start==prev else f"{start}..{prev}")
    return ",".join(parts)

def split_prefix_number(name: str) -> Tuple[str, int]:
    m = re.match(r"^([^\d]*?)(\d+)$", name)
    if not m: return None, None
    return m.group(1), int(m.group(2))

def is_hiddle_folder(name: str) -> bool:
    return name.endswith(".h")

def hiddle_label(name: str) -> str:
    return (name[:-2] + " [Hiddle]") if is_hiddle_folder(name) else name

# ======================= 진행률 관리자 (ETA 포함) =======================
class Progress:
    """
    전체 진행률(0~100)을 단계별 가중치로 집계:
      - COUNT (사전 카운팅): 10%
      - SCAN  (트리 스캔):   60%
      - BUILD (그룹/매트릭스): 20%
      - SAVE  (파일 저장):   10%
    SAVE 구간은 tick_save() 시점마다 ETA(EWMA)로 남은 시간을 추정해 메시지에 포함.
    """
    W_COUNT, W_SCAN, W_BUILD, W_SAVE = 0.10, 0.60, 0.20, 0.10

    def __init__(self, sink: Callable[[float, str], None]):
        self.sink = sink
        self.count_total = 1; self.count_done  = 0
        self.scan_total  = 1; self.scan_done   = 0
        self.build_total = 1; self.build_done  = 0
        self.save_total  = 1; self.save_done   = 0
        self._last_emit  = -1
        # ETA 추적
        self._save_t_last = None
        self._save_ewma = None  # 초/틱 (EWMA)

    def _fmt_eta(self, secs: float) -> str:
        if secs is None: return "ETA --:--"
        m, s = divmod(int(secs), 60)
        if m >= 100:
            return f"ETA {m:02d}m+"
        return f"ETA {m:02d}:{s:02d}"

    def set_totals(self, count_total=None, scan_total=None, build_total=None, save_total=None):
        if count_total is not None: self.count_total = max(1, count_total)
        if scan_total  is not None: self.scan_total  = max(1, scan_total)
        if build_total is not None: self.build_total = max(1, build_total)
        if save_total  is not None:
            self.save_total  = max(1, save_total)
            # SAVE 총량 갱신 시점마다 ETA 초기화
            self._save_t_last = None
            self._save_ewma = None
        self.emit("총 작업량 초기화")

    def tick_count(self, msg=""): self.count_done += 1; self.emit(msg)
    def tick_scan (self, msg=""): self.scan_done  += 1; self.emit(msg)
    def tick_build(self, msg=""): self.build_done += 1; self.emit(msg)

    def tick_save(self, msg=""):
        import time
        now = time.time()
        if self._save_t_last is not None:
            dt = max(1e-6, now - self._save_t_last)
            alpha = 0.3  # 최근치를 더 반영
            self._save_ewma = dt if self._save_ewma is None else (alpha*dt + (1-alpha)*self._save_ewma)
        self._save_t_last = now

        self.save_done += 1
        remain = max(0, self.save_total - self.save_done)
        eta = None if self._save_ewma is None else self._save_ewma * remain
        eta_msg = self._fmt_eta(eta)
        base = msg or "[SAVE]"
        self.emit(f"{base} ({self.save_done}/{self.save_total}) {eta_msg}")

    def emit(self, msg=""):
        p = (
            self.W_COUNT * (self.count_done / self.count_total) +
            self.W_SCAN  * (self.scan_done  / self.scan_total)  +
            self.W_BUILD * (self.build_done / self.build_total) +
            self.W_SAVE  * (self.save_done  / self.save_total)
        ) * 100.0
        pct = max(0.0, min(100.0, p))
        # SAVE 단계 전에는 95%를 넘지 않게 제한(완료처럼 보이는 것 방지)
        if self.save_done == 0:
            pct = min(pct, 95.0)
        if int(pct) != self._last_emit:
            self._last_emit = int(pct)
            self.sink(pct, msg)

# ======================= 0) 사전 카운팅(진행률 분모) =======================
def pre_count_inventory(root: str, check_stop: Callable[[], bool], progress: Progress, log=print):
    total_dirs = 0
    total_files = 0

    def listdir_safe(path: str):
        try:
            return list(os.scandir(path))
        except Exception:
            return []

    stack = [root]
    while stack:
        if check_stop(): raise RuntimeError("사용자 중지")
        d = stack.pop()
        total_dirs += 1
        progress.tick_count(f"[COUNT] {d}")
        for e in listdir_safe(d):
            if e.is_dir(follow_symlinks=False):
                if is_ignored_dir(e.name):
                    continue
                total_dirs += 1
                if is_hiddle_folder(e.name):
                    continue  # .h 내부 미탐색
                stack.append(e.path)
            else:
                ext = norm_ext(e.name)
                if ext in IGNORE_FILE_EXTS:
                    continue
                total_files += 1

    scan_total = total_dirs + total_files
    build_total = max(10, total_dirs // 2)
    progress.set_totals(count_total=max(1, total_dirs),
                        scan_total=max(1, scan_total),
                        build_total=build_total)
    log(f"[COUNT] 폴더 {total_dirs:,}개, 파일 {total_files:,}개(유효) 추정")

# ==================== 1) 전체 트리 스캔 ====================
def scan_tree(root: str, check_stop: Callable[[], bool], progress: Progress, log=print):
    files_rows: List[Dict[str, Any]] = []
    units_rows: List[Dict[str, Any]] = []

    def listdir_sorted(path: str):
        try:
            return sorted(os.scandir(path), key=lambda e: (not e.is_dir(), e.name.lower()))
        except Exception:
            return []

    def build_structure_signature(direct_files_norm: Tuple[str, ...], child_dirs: List[str]) -> str:
        nonnum = []
        num_map: Dict[str, List[int]] = {}
        for dn in child_dirs:
            pre, num = split_prefix_number(dn)
            if pre is not None and num is not None:
                num_map.setdefault(pre, []).append(num)
            else:
                nonnum.append(dn)
        nonnum_sorted = tuple(sorted(nonnum))
        num_parts = []
        for pre, nums in sorted(num_map.items()):
            num_parts.append(f"{pre}({compress_numbers(nums)})")
        files_sig = ",".join(direct_files_norm)
        return f"files=[{files_sig}]::dirs_nonnum=[{','.join(nonnum_sorted)}]::dirs_num=[{'|'.join(num_parts)}]"

    def dfs(path: str):
        if check_stop(): raise RuntimeError("사용자 중지")
        base = os.path.basename(path)
        direct_files = []
        child_dirs = []
        for ch in listdir_sorted(path):
            if ch.is_dir(follow_symlinks=False):
                if is_ignored_dir(ch.name):
                    continue
                child_dirs.append(ch.name)
            else:
                ext = norm_ext(ch.name)
                if ext in IGNORE_FILE_EXTS:
                    continue
                direct_files.append(normalize_filename(ch.name))
                st = safe_stat(ch.path)
                files_rows.append({
                    "Parent": path,
                    "Name": ch.name,
                    "NameNorm": normalize_filename(ch.name),
                    "Path": ch.path,
                    "Size(Bytes)": st.st_size if st else 0,
                    "Size(Human)": human_size(st.st_size) if st else ""
                })
                progress.tick_scan(f"[SCAN][FILE] {ch.path}")

        file_set = tuple(sorted(set(direct_files)))
        units_rows.append({
            "UnitPath": path,
            "ParentOfUnit": os.path.dirname(path),
            "UnitName": base,
            "FileSet": file_set,
            "FilesPerUnit": len(file_set),
            "NonNumChilds": tuple(sorted({n for n in child_dirs if split_prefix_number(n) == (None, None)})),
            "NumGroups": {}
        })
        # NumGroups
        pre_map: Dict[str, List[int]] = {}
        for dn in child_dirs:
            pre, num = split_prefix_number(dn)
            if pre is not None and num is not None:
                pre_map.setdefault(pre, []).append(num)
        units_rows[-1]["NumGroups"] = {k: sorted(v) for k, v in pre_map.items()}
        units_rows[-1]["StructureSig"] = build_structure_signature(file_set, child_dirs)
        progress.tick_scan(f"[SCAN][DIR ] {path}")

        if is_hiddle_folder(base):
            return
        for ch in listdir_sorted(path):
            if ch.is_dir(follow_symlinks=False):
                if is_ignored_dir(ch.name) or is_hiddle_folder(ch.name):
                    continue
                dfs(ch.path)

    dfs(root)
    return files_rows, units_rows

# ============== 2) 패턴 그룹 탐지 (구조 기반) ==============
def detect_groups_by_structure(root: str, check_stop: Callable[[], bool], progress: Progress, log=print):
    files_rows, units_rows = scan_tree(root, check_stop, progress, log)
    if check_stop(): raise RuntimeError("사용자 중지")

    df_units = pd.DataFrame(units_rows)
    if df_units.empty:
        return pd.DataFrame(), pd.DataFrame()

    df_units["MainRoot"] = os.path.basename(root)
    df_units[["Prefix","Number"]] = df_units["UnitName"].apply(lambda n: pd.Series(split_prefix_number(n)))

    grouped = (
        df_units.groupby(["MainRoot","ParentOfUnit","StructureSig","Prefix"], dropna=False)
        .agg(
            UnitCount=("UnitName","nunique"),
            FilesPerUnit=("FilesPerUnit","first"),
            UnitNames=("UnitName", lambda x: sorted(set(x))),
            Numbers=("Number", lambda x: sorted(int(n) for n in x if pd.notna(n))),
            Parents=("UnitPath", lambda x: sorted(set(x)))
        )
        .reset_index()
    )
    if check_stop(): raise RuntimeError("사용자 중지")

    def units_label(prefix, numbers, names):
        if pd.notna(prefix) and numbers:
            return f"{prefix}({compress_numbers(numbers)})"
        if len(names) <= 10:
            return ", ".join(names)
        return ", ".join(names[:10]) + f" … (+{len(names)-10})"

    grouped["UnitsLabel"] = grouped.apply(lambda r: units_label(r["Prefix"], r["Numbers"], r["UnitNames"]), axis=1)
    grouped["FileList"] = grouped["StructureSig"].apply(
        lambda s: s.split("::")[0].replace("files=[","").replace("]","")
    )
    grouped = grouped[grouped["UnitCount"] >= 2].copy()

    grouped = grouped.sort_values(["ParentOfUnit","UnitCount"], ascending=[True, False])
    grouped["PatternID"] = (
        grouped.groupby("MainRoot").cumcount().add(1).astype(str)
        .radd(grouped["MainRoot"] + " :: P")
    )
    progress.tick_build("[BUILD] 그룹 탐지 완료")
    return grouped.reset_index(drop=True), df_units.reset_index(drop=True)

# ============== 3) Collapsed Tree ==============
def make_collapsed_rows(root: str, grouped: pd.DataFrame, check_stop: Callable[[], bool],
                        progress: Progress, log=print) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if grouped is None or not isinstance(grouped, pd.DataFrame) or grouped.empty:
        grouped = pd.DataFrame()

    parent_to_groups: Dict[str, List[dict]] = {}
    if not grouped.empty:
        for _, g in grouped.iterrows():
            parent_to_groups.setdefault(g["ParentOfUnit"], []).append(g)

    grouped_member_paths: Set[str] = set()
    if not grouped.empty:
        for parents in grouped["Parents"]:
            grouped_member_paths.update(parents)

    def listdir_sorted(path: str):
        try:
            return sorted(os.scandir(path), key=lambda e: (not e.is_dir(), e.name.lower()))
        except Exception:
            return []

    def dfs(path: str, parts: List[str]):
        if check_stop(): raise RuntimeError("사용자 중지")
        base = os.path.basename(path)
        cur_parts = parts[:]
        cur_parts[-1] = hiddle_label(cur_parts[-1])

        rows.append({"Parts": cur_parts, "Kind": "Folder", "Name": cur_parts[-1], "Extra": ""})
        progress.tick_build(f"[BUILD][ROW ] {path}")
        if ROW_LIMIT and len(rows) >= ROW_LIMIT:
            return
        if is_hiddle_folder(base):
            return

        # 그룹 요약
        groups = parent_to_groups.get(path, [])
        if groups:
            for g in groups:
                label = f"[Group] {g['UnitsLabel']}  [{g['UnitCount']} folders]"
                rows.append({
                    "Parts": cur_parts[:] + [label],
                    "Kind": "Group",
                    "Name": g["UnitsLabel"],
                    "Extra": ""
                })
                progress.tick_build(f"[BUILD][GRUP] {label}")
                if ROW_LIMIT and len(rows) >= ROW_LIMIT:
                    return

        # 자식
        for ch in listdir_sorted(path):
            if ch.is_dir(follow_symlinks=False):
                if is_ignored_dir(ch.name):
                    continue
                if ch.path in grouped_member_paths:
                    continue
                if is_hiddle_folder(ch.name):
                    label = hiddle_label(ch.name)
                    rows.append({"Parts": cur_parts[:] + [label], "Kind": "Folder", "Name": label, "Extra": ""})
                    progress.tick_build(f"[BUILD][HIDE] {label}")
                    if ROW_LIMIT and len(rows) >= ROW_LIMIT:
                        return
                    continue
                dfs(ch.path, cur_parts + [ch.name])
                if ROW_LIMIT and len(rows) >= ROW_LIMIT:
                    return
            else:
                ext = norm_ext(ch.name)
                if ext in IGNORE_FILE_EXTS:
                    continue
                st = safe_stat(ch.path)
                rows.append({
                    "Parts": cur_parts[:] + [ch.name],
                    "Kind": "File",
                    "Name": ch.name,
                    "Extra": human_size(st.st_size) if st else ""
                })
                progress.tick_build(f"[BUILD][FILE] {ch.name}")
                if ROW_LIMIT and len(rows) >= ROW_LIMIT:
                    return

    dfs(root, [os.path.basename(root)])
    return rows

# ============== 4) 박스문자 트리 매트릭스 ==============
def build_box_matrix(df_paths: pd.DataFrame, level_cols: List[str]) -> pd.DataFrame:
    paths = [[r[c] for c in level_cols if r[c]] for _, r in df_paths.iterrows()]
    out = []
    for i, r in df_paths.iterrows():
        parts = paths[i]
        vis = {c: "" for c in level_cols}
        if not parts:
            out.append({**vis, "Kind": r["Kind"], "MainRoot": r["MainRoot"], "Extra": r["Extra"]})
            continue
        k = len(parts) - 1
        vis[level_cols[k]] = parts[-1]
        if k > 0:
            parent = tuple(parts[:-1])
            has_next = False
            for j in range(i+1, len(paths)):
                p2 = paths[j]
                if len(p2) < len(parent) or tuple(p2[:len(parent)]) != parent:
                    break
                if p2 != parts:
                    has_next = True
                    break
            connector = "├" if has_next else "└"
            vis[level_cols[k-1]] = connector
        for a in range(0, max(0, k-1)):
            ancestor = tuple(parts[:a+1])
            vertical = False
            for j in range(i+1, len(paths)):
                p2 = paths[j]
                if len(p2) <= a or p2[a] != parts[a]:
                    break
                if tuple(p2[:a+1]) == ancestor:
                    vertical = True
                    break
            if vertical:
                vis[level_cols[a]] = "│"
        out.append({**vis, "Kind": r["Kind"], "MainRoot": r["MainRoot"], "Extra": r["Extra"]})
    return pd.DataFrame(out, columns=level_cols + ["Kind","MainRoot","Extra"])

# ============== 5) 커넥터 가로선 패딩 ==============
def pad_connectors_to_width(ws, level_col_count: int, margin_chars: int = 1):
    from openpyxl.utils import get_column_letter
    target_len = {}
    for c in range(1, level_col_count + 1):
        letter = get_column_letter(c)
        width = ws.column_dimensions[letter].width or 10
        target_len[c] = max(1, int(round(width)) - margin_chars)
    for row in range(2, ws.max_row + 1):
        for c in range(1, level_col_count):
            cell = ws.cell(row=row, column=c)
            if cell.value is None: continue
            s = str(cell.value)
            if not s: continue
            if s and s[0] in ("├", "└"):
                head = s[0]
                cell.value = head + ("─" * max(0, target_len[c] - 1))

# ============== SAVE 단계 작업량 계산(진행률 세분화) ==============
def estimate_save_total(df_paths: pd.DataFrame, df_summary: pd.DataFrame, df_members: pd.DataFrame) -> int:
    # 기본: 시트 생성 4 + 스타일 4 + 패딩 1 + TXT 1
    total = 10
    n_rows = len(df_paths)
    n_tree_rows = n_rows
    n_sum = len(df_summary)
    n_mem = len(df_members)
    # 행 기반 추가: 500행/1000행 단위로 tick (너무 잦은 업데이트 방지)
    total += max(1, n_rows // 500)
    total += max(1, n_tree_rows // 500)
    total += max(1, max(1, n_sum // 1000))
    total += max(1, max(1, n_mem // 1000))
    # TXT도 행 기반
    total += max(1, n_rows // 2000)
    return total

# ============== 저장 (엑셀/텍스트) ==============
def save_outputs(root_dir: str, df_paths: pd.DataFrame, level_cols: List[str],
                 all_summaries, all_members, excel_basename: str,
                 check_stop: Callable[[], bool], progress: Progress, log=print):
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment

    excel_path = os.path.join(root_dir, excel_basename + ".xlsx")
    txt_collapsed_path = os.path.join(root_dir, excel_basename + "_collapsed.txt")

    df_summary = pd.concat(all_summaries, ignore_index=True) if all_summaries else pd.DataFrame(
        columns=["PatternID","MainRoot","ParentOfUnit","StructureSig","Prefix",
                 "UnitCount","FilesPerUnit","UnitNames","Numbers","Parents","UnitsLabel","FileList"]
    )
    df_members = pd.concat(all_members, ignore_index=True) if all_members else pd.DataFrame(
        columns=["PatternID","MainRoot","ParentPath","UnitName","FilesPerUnit","ParentOfUnit","UnitsLabel"]
    )

    # SAVE 총량을 실제 데이터 크기로 갱신
    progress.set_totals(save_total=estimate_save_total(df_paths, df_summary, df_members))
    # SAVE 시작 즉시 95% 탈출(시작 틱)
    progress.tick_save("[SAVE] 시작")

    # === Excel 저장 ===
    import openpyxl  # ensure packaged
    try:
        with pd.ExcelWriter(excel_path, engine="openpyxl") as w:
            # Tree_Collapsed
            df_excel = build_box_matrix(df_paths, level_cols)
            df_excel.to_excel(w, index=False, sheet_name="Tree_Collapsed")
            progress.tick_save("[SAVE] Tree_Collapsed 시트 작성")

            # Tree_View (행 많을 수 있으니 청크 로그/틱)
            tree_rows = []
            for idx, r in df_paths.iterrows():
                if check_stop(): raise RuntimeError("사용자 중지")
                parts = [r[c] for c in level_cols if r[c]]
                if not parts: continue
                if len(parts) == 1:
                    tree_rows.append({"Tree": parts[0], "Kind": r["Kind"], "Extra": r["Extra"]})
                else:
                    tree_rows.append({"Tree": ("    " * (len(parts)-1)) + "└── " + parts[-1],
                                      "Kind": r["Kind"], "Extra": r["Extra"]})
                if (idx+1) % 500 == 0:
                    progress.tick_save(f"[SAVE] Tree_View 구성 {idx+1:,}행")

            pd.DataFrame(tree_rows, columns=["Tree","Kind","Extra"]).to_excel(
                w, index=False, sheet_name="Tree_View")
            progress.tick_save("[SAVE] Tree_View 시트 작성")

            # PatternSummary
            if not df_summary.empty:
                df_summary[["PatternID","MainRoot","ParentOfUnit","UnitsLabel","UnitCount"]].to_excel(
                    w, index=False, sheet_name="PatternSummary")
            progress.tick_save("[SAVE] PatternSummary 시트 작성")

            # PatternMembers (많을 수 있음)
            if not df_members.empty:
                log(f"[SAVE] PatternMembers {len(df_members):,}행 작성")
                df_members.to_excel(w, index=False, sheet_name="PatternMembers")
            progress.tick_save("[SAVE] PatternMembers 시트 작성")

            # 스타일
            wb = w.book

            def style_ws(ws, wide_first=False, name=""):
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.font = Font(name=FONT_NAME, size=FONT_SIZE)
                        cell.alignment = Alignment(horizontal="left", vertical="center",
                                                   wrap_text=False, shrink_to_fit=False)
                for c in range(1, ws.max_column + 1):
                    letter = get_column_letter(c)
                    max_len = 0
                    for cell in ws[letter]:
                        if cell.value is not None:
                            max_len = max(max_len, len(str(cell.value)))
                    base = (LEVEL0_EXTRA_WIDTH if (c == 1 and wide_first) else DEFAULT_LEVEL_WIDTH)
                    width = max(base, max_len + 1)
                    width = min(EXCEL_MAX_COL_WIDTH, max(EXCEL_MIN_COL_WIDTH, width))
                    ws.column_dimensions[letter].width = width
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions
                progress.tick_save(f"[SAVE] 스타일 적용: {name}")

            if "Tree_Collapsed" in wb.sheetnames: style_ws(wb["Tree_Collapsed"], wide_first=True, name="Tree_Collapsed")
            if "Tree_View" in wb.sheetnames:      style_ws(wb["Tree_View"], wide_first=True, name="Tree_View")
            if "PatternSummary" in wb.sheetnames:  style_ws(wb["PatternSummary"], name="PatternSummary")
            if "PatternMembers" in wb.sheetnames:  style_ws(wb["PatternMembers"], name="PatternMembers")

            # ─ 패딩
            ws_tc = wb["Tree_Collapsed"]
            level_col_count = sum(1 for c in ws_tc[1] if str(c.value).startswith("Level"))
            pad_connectors_to_width(ws_tc, level_col_count, margin_chars=1)
            progress.tick_save("[SAVE] 커넥터 패딩 적용")

        progress.tick_save("[SAVE] 엑셀 파일 저장 완료")
    except Exception as e:
        # 부분 파일 제거
        try:
            if os.path.exists(excel_path): os.remove(excel_path)
        except Exception:
            pass
        log(f"[ERROR] Excel 저장 실패: {e}")
        raise

    # === TXT 저장 ===
    try:
        with open(txt_collapsed_path, "w", encoding="utf-8") as f:
            last_root = None
            for idx, r in df_paths.iterrows():
                if check_stop(): raise RuntimeError("사용자 중지")
                part_cols = [c for c in df_paths.columns if c.startswith("Level")]
                parts = [r[c] for c in part_cols if r[c]]
                if not parts:
                    if (idx+1) % 2000 == 0:
                        progress.tick_save(f"[SAVE] TXT 작성 {idx+1:,}행")
                    continue
                if parts[0] != last_root:
                    last_root = parts[0]
                    f.write("\n" + "="*80 + f"\n[{last_root}]\n" + "="*80 + "\n")
                    f.write(last_root + "\n")
                if len(parts) > 1:
                    indent = "    " * (len(parts)-1)
                    f.write(indent + "└── " + parts[-1] + "\n")
                if (idx+1) % 2000 == 0:
                    progress.tick_save(f"[SAVE] TXT 작성 {idx+1:,}행")
        progress.tick_save("[SAVE] TXT 저장 완료")
    except Exception as e:
        try:
            if os.path.exists(txt_collapsed_path): os.remove(txt_collapsed_path)
        except Exception:
            pass
        log(f"[ERROR] TXT 저장 실패: {e}")
        raise

    return excel_path, txt_collapsed_path

# ============== 메인 파이프라인 ==============
def run_backup_tree(root_dir: str, check_stop: Callable[[], bool],
                    emit_progress: Callable[[float, str], None], log=print):
    if not root_dir or not os.path.isdir(root_dir):
        raise ValueError("유효한 root_dir 폴더를 선택/입력하세요.")

    progress = Progress(emit_progress)
    log(f"[INIT] root_dir = {root_dir}")
    emit_progress(0.0, "초기화")

    # 0) 사전 카운트
    pre_count_inventory(root_dir, check_stop, progress, log)

    save_basename = f"_BackupTree_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    # Main roots
    main_roots = find_main_roots(root_dir)
    log(f"[INIT] main_roots: {', '.join(os.path.basename(m) for m in main_roots)}")

    all_rows: List[Dict[str, Any]] = []
    all_summaries: List[pd.DataFrame] = []
    all_members: List[pd.DataFrame] = []

    # 1~2) 스캔 + 그룹 + Collapsed
    for mr in main_roots:
        if check_stop(): raise RuntimeError("사용자 중지")
        log(f"[SCAN] {mr}")
        grouped, _ = detect_groups_by_structure(mr, check_stop, progress, log)
        rows = make_collapsed_rows(mr, grouped, check_stop, progress, log)
        for r in rows:
            r["MainRoot"] = os.path.basename(mr)
        all_rows.extend(rows)

        if isinstance(grouped, pd.DataFrame) and not grouped.empty:
            all_summaries.append(grouped.assign(MainRoot=os.path.basename(mr)))
            mem_rows = []
            for _, g in grouped.iterrows():
                for uname, pth in zip(g["UnitNames"], g["Parents"]):
                    mem_rows.append({
                        "PatternID": g["PatternID"],
                        "MainRoot": g["MainRoot"],
                        "ParentPath": pth,
                        "UnitName": uname,
                        "FilesPerUnit": g["FilesPerUnit"],
                        "UnitsLabel": g["UnitsLabel"],
                        "ParentOfUnit": g["ParentOfUnit"]
                    })
            all_members.append(pd.DataFrame(mem_rows))

    if check_stop(): raise RuntimeError("사용자 중지")
    if not all_rows:
        raise RuntimeError("스캔 결과가 없습니다.")

    # 3) 레벨 컬럼 구성
    df = pd.DataFrame(all_rows)
    max_depth = max(len(p) for p in df["Parts"])
    for i in range(max_depth):
        df[f"Level{i}"] = df["Parts"].apply(lambda p, i=i: p[i] if i < len(p) else "")
    level_cols = [f"Level{i}" for i in range(max_depth)]
    df_paths = df[level_cols + ["Kind","MainRoot","Extra"]].copy()
    progress.tick_build("[BUILD] 박스 매트릭스 준비 완료")

    # 4) 저장
    log("[SAVE] 파일 저장 시작")
    excel_path, txt_path = save_outputs(root_dir, df_paths, level_cols,
                                        all_summaries, all_members, save_basename,
                                        check_stop, progress, log)

    emit_progress(100.0, "완료")
    log("[DONE] 저장 완료")
    log(f"[DONE] Excel: {excel_path}")
    log(f"[DONE] TXT  : {txt_path}")
    return excel_path, txt_path

# ============================ GUI ============================
def launch_gui():
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox

    class App:
        def __init__(self, master):
            self.master = master
            master.title("Backup Tree (V17d)")
            master.geometry("860x560")

            # ---- 상태/스레드 ----
            self.q = queue.Queue()
            self.worker = None
            self._stop_flag = False

            # ---- 설정 로드 ----
            cfg = load_config()
            default_root = cfg.get("default_root") or os.path.expanduser("~")

            # ---- 상단: 경로 행 ----
            top = ttk.Frame(master)
            top.pack(fill="x", padx=10, pady=10)

            ttk.Label(top, text="루트 경로:").grid(row=0, column=0, sticky="w")
            self.var_root = tk.StringVar(value=default_root)
            self.ent_root = ttk.Entry(top, textvariable=self.var_root, width=80)
            self.ent_root.grid(row=0, column=1, sticky="we", padx=(6,6))
            top.columnconfigure(1, weight=1)

            self.btn_change = ttk.Button(top, text="경로 변경…", command=self.on_change_root)
            self.btn_change.grid(row=0, column=2, sticky="e")

            self.var_save_default = tk.BooleanVar(value=False)
            self.chk_default = ttk.Checkbutton(top, text="이 경로를 기본값으로 저장", variable=self.var_save_default)
            self.chk_default.grid(row=1, column=1, sticky="w", pady=(6,0))

            # ---- 중단/실행 ----
            mid = ttk.Frame(master)
            mid.pack(fill="x", padx=10, pady=(0,6))
            self.btn_run = ttk.Button(mid, text="실행", command=self.on_run)
            self.btn_run.pack(side="left")
            self.btn_stop = ttk.Button(mid, text="중지", command=self.on_stop, state="disabled")
            self.btn_stop.pack(side="left", padx=(8,0))

            # ---- 진행률 ----
            pr = ttk.Frame(master)
            pr.pack(fill="x", padx=10, pady=(0,6))
            self.var_pct = tk.StringVar(value="0%")
            ttk.Label(pr, text="진행률:").pack(side="left")
            self.lbl_pct = ttk.Label(pr, textvariable=self.var_pct, width=6)
            self.lbl_pct.pack(side="left", padx=(6,12))
            self.pb = ttk.Progressbar(pr, mode="determinate", maximum=100)
            self.pb.pack(fill="x", expand=True)

            # ---- 로그 ----
            frm_log = ttk.LabelFrame(master, text="실시간 로그")
            frm_log.pack(fill="both", expand=True, padx=10, pady=(0,10))
            self.txt = tk.Text(frm_log, height=20, wrap="word")
            self.txt.pack(fill="both", expand=True)

            # 폴링 루프 & 창 닫기
            master.protocol("WM_DELETE_WINDOW", self.on_close)
            self.master.after(75, self.poll_queue)

            self.log("[READY] 경로 확인 후 [실행]을 누르세요.")

        # ---------- 유틸 ----------
        def log(self, s: str):
            self.q.put(("log", s))

        def set_progress(self, pct: float, msg: str):
            self.q.put(("progress", (pct, msg)))

        def poll_queue(self):
            try:
                while True:
                    kind, payload = self.q.get_nowait()
                    if kind == "log":
                        self.txt.insert("end", payload + "\n")
                        self.txt.see("end")
                    elif kind == "progress":
                        pct, msg = payload
                        self.pb['value'] = pct
                        self.var_pct.set(f"{int(pct)}%")
                        if msg:
                            self.txt.insert("end", f"{int(pct)}% - {msg}\n")
                            self.txt.see("end")
                    elif kind == "done":
                        self.btn_run.config(state="normal")
                        self.btn_stop.config(state="disabled")
                        excel, txtp = payload
                        self.pb['value'] = 100
                        self.var_pct.set("100%")
                        messagebox.showinfo("완료", f"작업 완료!\n\nExcel:\n{excel}\n\nTXT:\n{txtp}")
                    elif kind == "error":
                        self.btn_run.config(state="normal")
                        self.btn_stop.config(state="disabled")
                        messagebox.showerror("오류", str(payload))
                    self.q.task_done()
            except queue.Empty:
                pass
            self.master.after(50, self.poll_queue)

        def on_change_root(self):
            from tkinter import filedialog
            d = filedialog.askdirectory(title="루트 경로 선택", initialdir=self.var_root.get() or os.path.expanduser("~"))
            if d:
                self.var_root.set(d)

        def on_run(self):
            root_dir = self.var_root.get().strip()
            if not root_dir or not os.path.isdir(root_dir):
                from tkinter import messagebox
                messagebox.showerror("유효하지 않은 경로", "존재하는 폴더를 선택하세요.")
                return
            # 기본값 저장 요청 시
            if self.var_save_default.get():
                cfg = load_config()
                cfg["default_root"] = root_dir
                save_config(cfg)
                self.log(f"[CONFIG] 기본 경로 저장: {root_dir}")

            self._stop_flag = False
            self.btn_run.config(state="disabled")
            self.btn_stop.config(state="normal")
            self.pb['value'] = 0
            self.var_pct.set("0%")
            self.txt.insert("end", "\n=== 작업 시작 ===\n")

            def job():
                try:
                    def check_stop():
                        return self._stop_flag
                    def emit_progress(pct, msg):
                        self.set_progress(pct, msg)
                    def gui_log(*args):
                        self.log(" ".join(str(a) for a in args))
                    excel_path, txt_path = run_backup_tree(root_dir, check_stop, emit_progress, log=gui_log)
                    self.q.put(("done", (excel_path, txt_path)))
                except Exception as e:
                    self.q.put(("error", e))

            self.worker = threading.Thread(target=job, daemon=True)
            self.worker.start()

        def on_stop(self):
            self._stop_flag = True
            self.btn_stop.config(state="disabled")
            self.log("[STOP] 중지 요청 → 가능한 지점에서 즉시 중단 및 부분 파일 정리")

        def on_close(self):
            # 안전 종료: 작업 중이면 stop flag, 잠시 대기 후 종료
            self._stop_flag = True
            if self.worker and self.worker.is_alive():
                try:
                    self.worker.join(timeout=2.0)
                except Exception:
                    pass
            try:
                self.master.destroy()
            finally:
                os._exit(0)  # 프로세스 완전 종료 보장

    # DPI 보정(Windows)
    try:
        if sys.platform.startswith("win"):
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass

    import tkinter as tk
    root = tk.Tk()
    App(root)
    root.mainloop()

# ========================== 실행 진입점 ==========================
if __name__ == "__main__":
    # CLI: 인자로 경로 주면 GUI 없이 바로 실행
    if len(sys.argv) > 1:
        target = sys.argv[1]
        def emit(p, m): print(f"{int(p):3d}% - {m}")
        def stop(): return False
        def lg(*a): print(" ".join(str(x) for x in a))
        run_backup_tree(target, stop, emit, log=lg)
    else:
        launch_gui()
